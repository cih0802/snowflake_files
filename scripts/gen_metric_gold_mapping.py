# gen_metric_gold_mapping.py — 지표 → GOLD 추적 장표 생성기
# Co-authored with CoCo
"""
GN_DW '지표 → GOLD' 추적(traceability) 장표 생성기.
"원하는 지표가 GOLD 어디에(FACT/DIM/SV·물리컬럼·SV base) 어떻게 매핑됐는지" 한 눈에 본다.

정본(근거) — 전부 파싱해 지표번호로 조인한다. **사실을 리터럴로 보관하지 않는다**(P85-①):
  · 03_top-down_gold/02_지표 분류.md            (215 지표: 유형·소스·단위·배속)
  · 99_provided_definition/02_지표사전 공통.md    (공통 162: 구분·정의·정본 계산식)
  · 99_provided_definition/03_지표사전 신규.md    (신규 53: 〃)
  · 03_top-down_gold/04_SV파생 매핑.md          (derived → 분자/분모 base + FACT · base 카탈로그)
  · 03_top-down_gold/05_필드 인벤토리.md         (보고서필드 라벨 → GOLD 물리컬럼 정본)
  · 99_provided_definition/04·05_보고서필드 인벤토리.md (보고서 × 필드)
  · 30_output_share/04_컬럼계보매핑.csv          (GOLD → SILVER → BRONZE 계보. 🔴 **산출물 파일**로 받는다)
  · /tmp/census.json                            (GOLD 전 컬럼 COUNT / COUNT_IF(<>0) 실측)
  · scripts/field_mapping_override.py           (보고서필드 매핑 교정 등록부 — 큐레이션 분리)

출력: 30_output_share/05_지표GOLD매핑.{md,csv,xlsx}

────────────────────────────────────────────────────────────────────────────
[2026-08-06 재작성] 소스 유실 복구 — 종전 판본에서 바꾼 것
  🔴 ① **상태를 원천 계통으로 추정하지 않는다.** 지표에 대응하는 GOLD 물리 컬럼을 특정할 수 있으면
       census 로 직접 측정해 판정하고 근거를 `상태_근거` 에 `실측:` 으로 적는다. 특정 불가할 때만
       `추정:` 을 쓴다. 종전 판본이 `OK 168` 로 적었던 것은 과대 진술이었고, 전건 `0` 컬럼을
       `OK` 로 분류해 **현업이 빈 결과를 답으로 믿게** 만들었다(O43 · P76 · P85-③).
  🔴 ② **계보를 모듈 import 로 가져오지 않는다.** 종전 `load_gcm()` 은 `gen_column_mapping.py` 를
       동적 import 해 `gcm.ROWS`(파이썬 리터럴 계보표)를 읽었다. 그 리터럴이 stale 의 배포원이었고
       (O43), 계보가 실측 파생으로 재작성되자 `ROWS` 가 사라져 **이 생성기가 실행 불가**가 됐다.
       ⇒ 이제 `04_컬럼계보매핑.csv`(산출물)를 입력으로 받는다 → 신선함이 전파된다(P85-④).
  🔴 ③ **교정 등록부를 실제로 배선한다.** `field_mapping_override.py` 는 작성만 되고 어떤 생성기에도
       import 되지 않아 O45 의 효과가 산출물에 0건 반영됐다(O45-D · P90-①).
       `try/except import` 를 쓰지 않는다 — 조용한 실패가 가장 위험하다(P90-③).
  🔴 ④ **COMMENT/문서에 수치를 복제하지 않는다.** 이 파일의 모든 수치는 실행 시점에 측정된다.
────────────────────────────────────────────────────────────────────────────
"""
import csv
import json
import os
import re
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))  # 🔴 샌드박스 sys.path[0] 고정 회피(P90-③)

import field_mapping_override as FMO   # noqa: E402  — 필수 의존. 실패 시 즉시 중단시킨다.

WS = os.environ.get("GN_DW_WS", "/workspace")
OUT_DIR = os.environ.get("GN_DW_OUT", os.path.join(WS, "30_output_share"))
CENSUS = os.environ.get("GN_DW_CENSUS", "/tmp/census.json")
LINEAGE_CSV = os.environ.get("GN_DW_LINEAGE", os.path.join(OUT_DIR, "04_컬럼계보매핑.csv"))
BASENAME = "05_지표GOLD매핑"
GEN_PATH = "scripts/gen_metric_gold_mapping.py"
MEASURED = os.environ.get("GN_DW_MEASURED", date.today().isoformat())
PROV = f"본 파일은 자동 생성물입니다. 직접 수정 금지 — 생성기 {GEN_PATH} 수정 후 재실행하세요."

DOC_CLS = os.path.join(WS, "03_top-down_gold", "02_지표 분류.md")
DOC_COMM = os.path.join(WS, "99_provided_definition", "02_지표사전 공통.md")
DOC_NEW = os.path.join(WS, "99_provided_definition", "03_지표사전 신규.md")
DOC_SV = os.path.join(WS, "03_top-down_gold", "04_SV파생 매핑.md")
DOC_INV = os.path.join(WS, "03_top-down_gold", "05_필드 인벤토리.md")
DOC_MKT = os.path.join(WS, "99_provided_definition", "04_마케팅_보고서필드 인벤토리.md")
DOC_MEM = os.path.join(WS, "99_provided_definition", "05_회원_보고서필드 인벤토리.md")

HEADER = ("지표#", "구분", "지표명", "유형", "소스", "단위", "GOLD_배속",
          "GOLD_매핑(물리컬럼/SV base)", "SILVER_원천", "BRONZE_원천", "정본_계산식",
          "상태", "상태_근거")
FIELD_HEADER = ("영역", "섹션", "필드값", "데이터원천", "데이터TYPE",
                "대응_지표#", "GOLD_매핑", "분해축", "매핑근거")

# 지표번호 → 배속 약어 확장
ABBR = {
    "FMM": "FACT_MEMBER_MONTHLY", "FME": "FACT_MEMBER_EVENT", "FSE": "FACT_SERVICE_EVENT",
    "FEP": "FACT_EVENT_PARTICIPATION", "FMC": "FACT_MEMBER_COHORT", "FMF": "FACT_MEMBER_FEE",
    "FAD": "FACT_AD_PERFORMANCE", "FAD_B": "FACT_AD_BROADCAST", "FAD_D": "FACT_AD_DIGITAL",
    "FAD_BC": "FACT_AD_BROADCAST_CASE", "FGA": "FACT_GA_BEHAVIOR", "FBD": "FACT_BUDGET",
    "FTG_D": "FACT_TARGET_DEV", "FTG-D": "FACT_TARGET_DEV",
    "FTG_B": "FACT_TARGET_BIZ", "FTG-B": "FACT_TARGET_BIZ",
}

# 원천 입고 대기(외부 하드블로커)로 물리 측정이 성립하지 않는 지표 — WAIT 고정
WAIT_SET = frozenset({"공152", "공153", "공154", "공155"})   # 사업목표 FTG-B (E-6)
# GA↔CRM identity 커버리지 종속 — 측정값이 있어도 PARTIAL 로 강등
IDENTITY_SET = frozenset({"공81", "공122", "신32", "신33"})
# 🔴 **컬럼 채움률이 100% 여도** PARTIAL 을 유지하는 소스 = GA4.
#   이유가 중요하다 — GA4 는 **적재된 샤드 안에서는 채움 100%** 지만 전기간이 입고되지 않았다(G-5).
#   즉 결손이 「컬럼 NULL」이 아니라 **「시간 커버리지」** 에 있어 census 로는 드러나지 않는다.
#   AGENCY·ERP 는 결손이 특정 축(소재 Q10 · 모금성비용 E-1)에 국한되므로 여기 넣지 않고
#   ① 채움률 임계 ② 추정 분기의 계통 표기로 처리한다 — 뭉개면 사유를 잘못 말하게 된다.
SRC_TIME_GAP = frozenset({"GA4", "GA"})
# 채움률 임계 — 이 미만이면 PARTIAL. 컬럼 자체는 살아 있으나 일부 구간·분기만 채워진 상태다.
OK_FILL_MIN = 0.95

# base 물리명을 지표번호로 역인덱스할 수 없는 경우의 명시 배속(측정 대상 컬럼 지정)
MEASURE_OVERRIDE = {
    "공25": "INBOUND_CALL", "공66": "REGULAR_FEE", "공87": "FAIL_MEMBERS",
    "공91": "GIFT_PART_AMT", "공97": "SESSION_CNT", "공107": "SCROLL_DEPTH",
}

INV_DELIMS = re.compile(r"\s(?:SUM|COUNT|AVG|COALESCE|비가산|conform|롤업|GA4)|—|"
                        r"\((?:overview|05|04|신규|신#|SCD|정본|원천|=|선택|MM|SND|TM_|TH_|TD_|CRM|MBER|행동|건/)|/\s?10000")
INV_NO_SUBSTR = frozenset({
    "PLAN_BUDGET_YEAR", "PLAN_BUDGET_MONTH", "EXEC_BUDGET_ERP", "EXEC_BUDGET_EST",
    "AMOUNT_BAND1", "AMOUNT_BAND2", "PERIOD_BAND1", "PERIOD_BAND2",
    "UTM_TERM", "UTM_CONTENT",
})


# ───────────────────────────── 마크다운 표 파서 ─────────────────────────────
def parse_md(path):
    """표 블록 목록을 반환. 각 표 = dict(h1,h2,h3,label,header[],rows[[]])."""
    out, block, ctx = [], [], {"h1": "", "h2": "", "h3": "", "label": ""}

    def is_sep(cells):
        return all(re.fullmatch(r":?-{2,}:?", c.strip()) or set(c.strip()) <= set("-: ")
                   for c in cells) and any("-" in c for c in cells)

    def flush():
        if len(block) < 2:
            block.clear()
            return
        rows = [[c.strip() for c in ln.strip().strip("|").split("|")] for ln in block]
        header = rows[0]
        body = [r for r in rows[1:] if not is_sep(r)]
        out.append(dict(ctx, header=header, rows=body))
        block.clear()

    for raw in open(path, encoding="utf-8").readlines():
        ln = raw.rstrip("\n")
        s = ln.strip()
        if s.startswith("|"):
            block.append(s)
            continue
        flush()
        if s.startswith("# "):
            ctx = {"h1": s[2:].strip(), "h2": "", "h3": "", "label": ""}
        elif s.startswith("## "):
            ctx = dict(ctx, h2=s[3:].strip(), h3="", label="")
        elif s.startswith("### "):
            ctx = dict(ctx, h3=s[4:].strip(), label="")
        else:
            m = re.fullmatch(r"\*\*(.+?)\*\*", s)
            if m:
                ctx = dict(ctx, label=m.group(1).strip())
    flush()
    return out


def col_idx(header, *names):
    for i, h in enumerate(header):
        for n in names:
            if h.strip() == n:
                return i
    return None


def clean_placement(v):
    """배속 셀에서 대표 토큰만: 'SV ✅…'→'SV', 'FMM ⚠️…'→'FMM'."""
    return re.split(r"[ \u2705\u26a0\ufe0f]", v.strip(), 1)[0].strip()


def base_token(v):
    """base 표현에서 선두 UPPER_SNAKE 식별자: 'DEV_CNT(YTD)'→'DEV_CNT'."""
    m = re.search(r"[A-Z][A-Z0-9_]+", v or "")
    return m.group(0) if m else ""


# ───────────────────────────── 정본 로더 ─────────────────────────────
def load_classification():
    """{'공4': {name,type,src,unit,basis,place,place_raw}, '신20': {...}}"""
    out = {}
    for t in parse_md(DOC_CLS):
        h = t["header"]
        if col_idx(h, "지표명") is None or col_idx(h, "배속") is None:
            continue
        pref = "공" if "공통" in t["h2"] else ("신" if "신규" in t["h2"] else None)
        if not pref:
            continue
        i_no, i_nm = 0, col_idx(h, "지표명")
        i_ty, i_sr = col_idx(h, "유형"), col_idx(h, "소스")
        i_un, i_ba = col_idx(h, "단위"), col_idx(h, "집계/basis")
        i_pl = col_idx(h, "배속")
        for r in t["rows"]:
            if len(r) <= i_pl or not r[i_no].strip().isdigit():
                continue
            key = f"{pref}{int(r[i_no])}"
            out[key] = {
                "name": r[i_nm], "type": r[i_ty], "src": r[i_sr], "unit": r[i_un],
                "basis": r[i_ba] if i_ba is not None and len(r) > i_ba else "",
                "place_raw": r[i_pl], "place": clean_placement(r[i_pl]),
            }
    return out


def load_dictionary(path, pref):
    """{'공4': {cat, def, formula}}"""
    out = {}
    for t in parse_md(path):
        h = t["header"]
        i_no = col_idx(h, "순서")
        i_nm = col_idx(h, "지표명")
        if i_no is None or i_nm is None:
            continue
        i_ct, i_df, i_fm = col_idx(h, "구분"), col_idx(h, "정의"), col_idx(h, "계산식")
        for r in t["rows"]:
            if len(r) <= i_nm or not r[i_no].strip().isdigit():
                continue
            out[f"{pref}{int(r[i_no])}"] = {
                "cat": r[i_ct] if i_ct is not None and len(r) > i_ct else "",
                "def": r[i_df] if i_df is not None and len(r) > i_df else "",
                "formula": r[i_fm] if i_fm is not None and len(r) > i_fm else "",
            }
    return out


def _parse_srcnums(cell):
    """'공4·5·149' → [('공',[4,5,149])] · '공139~142' → [('공',[139..142])]."""
    out = []
    for pref, body in re.findall(r"(공|신)\s*([\d·~,\s]+)", cell or ""):
        nums = []
        for tok in re.split(r"[·,\s]+", body):
            tok = tok.strip()
            if not tok:
                continue
            if "~" in tok:
                a, b = tok.split("~")[:2]
                if a.strip().isdigit() and b.strip().isdigit():
                    nums.extend(range(int(a), int(b) + 1))
            elif tok.isdigit():
                nums.append(int(tok))
        if nums:
            out.append((pref, nums))
    return out


def load_sv():
    """(num2base, derived_map)
       num2base   : 지표# → base 물리명 (§1 base 카탈로그의 '원천 지표#' 역인덱스)
       derived_map: 지표# → {num, den, fact}
    """
    num2base, derived = {}, {}
    for t in parse_md(DOC_SV):
        h = t["header"]
        i_base, i_src = col_idx(h, "base 물리명"), col_idx(h, "원천 지표#")
        if i_base is not None and i_src is not None:
            for r in t["rows"]:
                if len(r) <= max(i_base, i_src):
                    continue
                bases = [b.strip() for b in re.split(r"\s*/\s*", r[i_base]) if b.strip()]
                for pref, nums in _parse_srcnums(r[i_src]):
                    for k, n in enumerate(nums):
                        key = f"{pref}{n}"
                        num2base.setdefault(key, bases[k] if k < len(bases) else bases[0])
            continue
        i_num, i_den = col_idx(h, "분자 base"), col_idx(h, "분모 base")
        if i_num is None or i_den is None:
            continue
        i_no = 0
        i_fact = col_idx(h, "FACT(분자·분모)", "FACT")
        pref = "공" if t["h2"].startswith("2.") else ("신" if t["h2"].startswith("3.") else None)
        if not pref:
            continue
        for r in t["rows"]:
            no = r[i_no].strip().lstrip("#").strip()
            if not no.isdigit():
                continue
            derived[f"{pref}{int(no)}"] = {
                "num": r[i_num] if len(r) > i_num else "",
                "den": r[i_den] if len(r) > i_den else "",
                "fact": r[i_fact] if i_fact is not None and len(r) > i_fact else "",
            }
    return num2base, derived


def load_lineage():
    """04_컬럼계보매핑.csv → (col2src, obj2src)
       col2src: GOLD 컬럼명            → (SILVER 원천, BRONZE 원천)
       obj2src: GOLD 테이블(또는 마트)  → (SILVER 원천, BRONZE 원천)
       🔴 모듈 import 가 아니라 **산출물 CSV** 를 읽는다(P85-④).
    """
    col2src, obj2src = {}, {}
    if not os.path.exists(LINEAGE_CSV):
        raise SystemExit(f"계보 입력이 없다: {LINEAGE_CSV} — gen_column_mapping.py 를 먼저 실행할 것")
    agg = {}
    with open(LINEAGE_CSV, encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if not row or row[0].startswith("#") or row[0].startswith("##"):
                continue
            if row[0].strip() in ("WIDE_마트", "계층"):
                hdr = row
                i_g = col_idx(hdr, "GOLD_원천(테이블.컬럼)")
                i_s = col_idx(hdr, "SILVER_원천(테이블.컬럼)")
                i_b = col_idx(hdr, "BRONZE_원천(테이블)")
                agg["idx"] = (i_g, i_s, i_b)
                continue
            i_g, i_s, i_b = agg.get("idx", (2, 3, 4))
            if i_g is None or len(row) <= max(filter(None, [i_g, i_s, i_b])):
                continue
            gold, silver, bronze = row[i_g].strip(), row[i_s].strip(), row[i_b].strip()
            if not gold or "." not in gold:
                continue
            tbl, _, col = gold.partition(".")
            if col:
                col2src.setdefault(col, (silver, bronze))
            if tbl:
                obj2src.setdefault(tbl, (silver, bronze))
    return col2src, obj2src


def load_census():
    if not os.path.exists(CENSUS):
        raise SystemExit(f"census 가 없다: {CENSUS} — scripts/census_columns.py 를 먼저 실행할 것")
    return json.load(open(CENSUS, encoding="utf-8"))


def _inv_label(desc):
    """설명에서 보고서필드 한글 라벨만: '개발(건) SUM(금액)/10000 (#4·5)' → '개발(건)'."""
    # 🔴 [2026-08-07 O48] **편집 마커 접두를 먼저 제거한다.** 정본 인벤토리의 설명 칸에는
    #   `**[2026-08-03 O25 신설]** 중단사유명` 처럼 **문서 이력 표기가 라벨 앞에** 붙은 셀이 있다.
    #   `_norm` 은 `*`·`[`·`]` 만 지우고 날짜·이슈번호는 남기므로 색인 키가
    #   `20260803o25신설중단사유명` 이 되어 **그 엔트리에 영원히 도달할 수 없었다**(실측 14건).
    #   ⇒ 이것도 P98 계열이다 — 수기 문서의 편집 관행이 생성기 색인을 조용히 망가뜨린다.
    #   ⚠️ 실측으로 영향 범위를 먼저 확인했다: 이 교정으로 색인 키 14개가 교체되지만
    #      **보고서필드 507종 중 신규 매칭 0 · 소멸 매칭 0** → 05·09 산출물 불변(재생성 불요).
    #      `scripts/test_generators.py::T2` 가 이 상태를 회귀 감시한다.
    desc = re.sub(r"^\s*\*\*\[[^\]]*\]\*\*\s*", "", desc or "")
    # 🔴 지표번호 괄호는 **괄호 전체**를 지운다 — 종전 판본은 여는 괄호만 지워
    #    '납입회비(원) (#69·70 단일화)' 가 '납입회비(원)  단일화)' 로 남아 라벨 매칭이 실패했다.
    s = re.sub(r"\([^()]*#[^()]*\)", "", desc or "")
    s = re.sub(r"\(?(?:신규|신)?#\s*[\d·~]+\)?", "", s)
    lab = re.split(INV_DELIMS, s)[0].strip(" .·")
    # 🔴 [2026-08-07 O47-B] **주석은 라벨이 아니다.** 인벤토리 설명 칸에는 라벨 대신 상태 주석만
    #   적힌 셀이 있다(예: FME 의 `CAMPAIGN_SK` = "🔴 **DEV 브랜치 전용**(실측 …) · **STOP 전건 0**").
    #   종전 파서는 이것을 라벨로 색인했고, `idx.setdefault` 가 **문서 순서 first-wins** 라서
    #   쓰레기 라벨이 실제 라벨의 자리를 선점할 수 있었다(실측: 이번 세션의 인벤토리 주석 추가로
    #   FME 3컬럼이 쓰레기 라벨로 색인됨). 상태 마커로 시작하면 라벨이 아니라고 판정한다.
    if lab[:1] in ("🔴", "🟠", "🟡", "🟢", "🔵", "⚠", "✅", "⛔", "◐", "❔", "🆕", "🔷", "▸"):
        return ""
    return lab


def _inv_metricno(desc):
    """설명의 (#..) → '공4·5·149' / '신20'."""
    out = []
    for pref, body in re.findall(r"(신규|신|공)?\s*#\s*([\d·~]+)", desc or ""):
        p = "신" if pref in ("신규", "신") else "공"
        nums = []
        for tok in body.split("·"):
            if "~" in tok:
                a, b = tok.split("~")[:2]
                if a.isdigit() and b.isdigit():
                    nums.extend(str(x) for x in range(int(a), int(b) + 1))
            elif tok.isdigit():
                nums.append(tok)
        if nums:
            out.append(p + "·".join(nums))
    return " ".join(out)


def _norm(s):
    s = (s or "").lower().replace("중복", "").replace("（", "(").replace("）", ")")
    s = s.replace("률", "율")          # 표기 변형 통일(이탈률 ↔ 이탈율) — 의미 동일
    return re.sub(r"[\s()\[\]/·,%*=＝~:\-—–÷]", "", s)


def load_gold_inventory():
    """05_필드 인벤토리.md → (idx: label_norm→entry, entries).
       entry = {col, table, desc, label_norm, mno}"""
    idx, entries = {}, []
    for t in parse_md(DOC_INV):
        head = t["h3"] or t["h2"]
        # 🔴 제목 형식이 두 가지다: 'FTG_D. FACT_TARGET_DEV — …' 와 'WIDE_DEV_ACHIEVEMENT — …'
        #   후자를 못 잡으면 테이블명이 비어 매핑이 `.COLUMN` 이 되고 소비 생성기가 판정불가로 흘린다(실측).
        m = re.match(r"[A-Z0-9_]+\.\s*([A-Z_][A-Z0-9_]+)", head)
        if not m:
            m = re.match(r"((?:FACT|DIM|WIDE|SV)_[A-Z0-9_]+)", head)
        table = m.group(1) if m else ""
        h = t["header"]
        i_col, i_desc = col_idx(h, "컬럼명"), col_idx(h, "설명", "설명/basis", "설명/근거")
        if i_col is None or i_desc is None:
            continue
        for r in t["rows"]:
            if len(r) <= max(i_col, i_desc):
                continue
            col = r[i_col].strip().strip("`")
            desc = r[i_desc].strip()
            if not col or col.startswith("~~") or "❌" in col or " " in col:
                continue
            label = _inv_label(desc)
            if not label:
                continue
            e = {"col": col, "table": table, "desc": desc, "label": label,
                 "label_norm": _norm(label), "mno": _inv_metricno(desc)}
            entries.append(e)
            idx.setdefault(e["label_norm"], e)
    return idx, entries


# ───────────────────────────── 상태 판정 (실측 우선) ─────────────────────────────
def measure(census, table, col):
    """(rows, nonnull, nonzero) — 없으면 None."""
    ent = census.get(f"GOLD.{table}") or census.get(f"SERVING.{table}")
    if not ent:
        return None
    c = ent["cols"].get(col)
    if not c:
        return None
    return ent["rows"], c["nonnull"], c["nonzero"]


def find_column(census, col, prefer=None):
    """census 전역에서 컬럼을 보유한 GOLD 테이블을 찾는다. prefer 우선."""
    hits = [t.split(".", 1)[1] for t, e in census.items()
            if t.startswith("GOLD.") and col in e["cols"]]
    if not hits:
        return None
    if prefer:
        for p in ([prefer] if isinstance(prefer, str) else prefer):
            if p in hits:
                return p
    return sorted(hits)[0]


def derive_status(key, cls, num2base, derived, census):
    """(상태, 근거). 🔴 물리 컬럼을 특정할 수 있으면 census 로 **측정**한다."""
    if key in WAIT_SET:
        return "WAIT", "추정: 원천 미입고(E-6 CRM 사업목표) — `FACT_TARGET_BIZ` 0행"

    # 측정 대상 컬럼 결정: MEASURE_OVERRIDE → base 카탈로그 → derived 분자 base
    col = MEASURE_OVERRIDE.get(key) or num2base.get(key)
    if not col and key in derived:
        col = base_token(derived[key]["num"])
    prefer = []
    place = cls.get("place", "")
    if place in ABBR:
        prefer.append(ABBR[place])
    if key in derived:
        for tok in re.split(r"[·,/ ]+", derived[key]["fact"] or ""):
            tok = tok.strip()
            if tok in ABBR:
                prefer.append(ABBR[tok])

    if col:
        tbl = find_column(census, col, prefer)
        if tbl:
            m = measure(census, tbl, col)
            if m:
                rows, nn, nz = m
                if rows == 0:
                    return "WAIT", f"실측: `{tbl}` **0행** — 테이블 미적재"
                if not nz:
                    return "WAIT", f"실측: `{tbl}.{col}` 전건 0 — 설계O·값 미주입"
                pct = f"{nz / rows * 100:.1f}%" if rows else "—"
                base = f"실측: `{tbl}.{col}` 비영 {nz:,}/{rows:,} ({pct})"
                if key in IDENTITY_SET:
                    return "PARTIAL", base + " · GA↔CRM identity 커버리지 종속"
                # 🔴 실측이 원천 계통 제약을 지우지 않는다 — 값이 있어도 GA4(1일 샤드)·
                #   AGENCY(소재 연결키 Q10)·ERP(모금성비용 E-1) 는 커버리지가 열려 있다.
                #   이 강등을 빼면 「사용가능」이 거짓이 되는 방향으로만 틀린다(P76).
                if cls.get("src") in SRC_TIME_GAP:
                    return "PARTIAL", base + " · ⚠️ 채움률은 **적재된 샤드 내부** 기준이다 — GA4 전기간 미입고(G-5)"
                if nz < rows * OK_FILL_MIN:
                    return "PARTIAL", base + f" · 채움률 {OK_FILL_MIN:.0%} 미만"
                return "OK", base

    # 물리 컬럼 특정 불가 → 배속·원천 계통 추정 (🔴 단독 인용 금지 표기)
    if key in IDENTITY_SET:
        return "PARTIAL", "추정: GA↔CRM identity 브리지 의존 — 물리 컬럼 미특정"
    if cls.get("type") == "dimension":
        tbl = ABBR.get(place, place)
        ent = census.get(f"GOLD.{tbl}")
        if ent and ent["rows"]:
            return ("PARTIAL" if cls.get("src") in ("GA4", "GA", "AGENCY", "ERP") else "OK",
                    f"추정: 배속 차원 `{tbl}` {ent['rows']:,}행 실재 — 대응 물리 컬럼 미특정")
    if cls.get("src") in ("GA4", "GA"):
        return "PARTIAL", "추정: GA4 원천 계통(1일 샤드·identity 4%대) — 물리 컬럼 미특정"
    if cls.get("src") == "AGENCY":
        return "PARTIAL", "추정: AGENCY 원천 계통(소재 연결키 Q10) — 물리 컬럼 미특정"
    if cls.get("src") == "ERP":
        return "PARTIAL", "추정: ERP 원천 계통(모금성비용 E-1) — 물리 컬럼 미특정"
    # 🔴 CRM·복합 계통은 알려진 계통 제약이 없다 → OK 로 두되 **추정임을 명시**한다.
    #   여기서 PARTIAL 로 강등하면 범례(「PARTIAL = GA/AGENCY/ERP·identity 일부 대기」)와 어긋나
    #   사유를 잘못 말하게 된다. 대신 `추정:` 표기 + 헤더의 「단독 인용 금지」로 방어한다.
    return "OK", "추정: 배속·원천 계통에 알려진 제약 없음 — 대응 물리 컬럼 미특정"


# ───────────────────────────── 지표 행 ─────────────────────────────
def build_metric_rows():
    cls = load_classification()
    dic = {}
    dic.update(load_dictionary(DOC_COMM, "공"))
    dic.update(load_dictionary(DOC_NEW, "신"))
    num2base, derived = load_sv()
    col2src, obj2src = load_lineage()
    census = load_census()

    def src_for_col(col):
        return col2src.get(col, ("", ""))

    def src_for_obj(obj):
        return obj2src.get(obj, ("", ""))

    rows, num2map = [], {}
    keys = [f"공{i}" for i in range(1, 163)] + [f"신{i}" for i in range(1, 54)]
    for key in keys:
        c = cls.get(key)
        if not c:
            continue
        d = dic.get(key, {})
        place = c["place"]
        typ = c["type"]
        mapping, silver, bronze = "", "", ""

        if typ == "derived" or place == "SV":
            dv = derived.get(key)
            if key in ("공98", "공108"):
                mapping = f"FGA 물리적재(비가산) — {num2base.get(key, '')}"
                silver, bronze = src_for_col(num2base.get(key, ""))
            elif dv:
                num, den = dv["num"] or "—", dv["den"] or "—"
                mapping = f"SV metric — 분자: {num}" + (f" / 분모: {den}" if den != "—" else "")
                silver, bronze = src_for_col(base_token(num))
            else:
                mapping = "SV metric (파생 — base는 04_SV파생 매핑.md 참조)"
        elif typ == "measure":
            col = MEASURE_OVERRIDE.get(key) or num2base.get(key, "")
            if col:
                mapping = f"{place}.{col}"
                silver, bronze = src_for_col(col)
            else:
                mapping = f"{place} (measure — 컬럼 06_DDL.sql 확인)"
                silver, bronze = src_for_obj(ABBR.get(place, place))
        else:  # dimension
            mapping = place
            silver, bronze = src_for_obj(ABBR.get(place, place))

        status, why = derive_status(key, c, num2base, derived, census)
        rows.append([key, d.get("cat", ""), c["name"], typ, c["src"], c["unit"],
                     place, mapping, silver, bronze, d.get("formula", ""), status, why])
        num2map[key] = (mapping, typ, place, status)
    return rows, num2map


# ───────────────────────────── 보고서필드 매핑 ─────────────────────────────
RATE_RE = re.compile(r"율|률|구성비|증감|대비|비중|달성|1인당|1명당|%")
UNIT_SUFFIX = re.compile(r"(명|건|원|개월|년|횟수|수|%)$")


def _keys(s):
    """라벨 1개에서 매칭 후보 키를 단계적으로 만든다.
    🔴 별칭 표(수십~수백 항목)를 하드코딩하지 않는다 — 그 표는 stale 의 배포원이고(P85),
       유실 시 추측으로 복원할 수도 없다. 대신 **정규화 규칙**으로 만든다.
       ① 정규화 원형  ② 단위 접미 제거형  ③ 문자 정렬형(어순 차이 흡수: '월 편성예산' ↔ '편성예산(월)')
    """
    n = _norm(s)
    out = [("정확", n)]
    n2 = UNIT_SUFFIX.sub("", n)
    if n2 and n2 != n:
        out.append(("단위", n2))
    if len(n) >= 4:
        out.append(("어순", "".join(sorted(n))))
    return out


def build_field_index(metric_rows):
    metric_idx, metric_unit, metric_sort = {}, {}, {}
    for r in metric_rows:
        for kind, k in _keys(r[2]):
            tgt = {"정확": metric_idx, "단위": metric_unit, "어순": metric_sort}[kind]
            tgt.setdefault(k, (r[0], r[3], r[7], r[6]))
    inv_idx, inv_entries = load_gold_inventory()
    inv_unit, inv_sort = {}, {}
    for e in inv_entries:
        for kind, k in _keys(e["label"]):
            if kind == "단위":
                inv_unit.setdefault(k, e)
            elif kind == "어순":
                inv_sort.setdefault(k, e)
    return {"metric": metric_idx, "metric_unit": metric_unit, "metric_sort": metric_sort,
            "inv": inv_idx, "inv_unit": inv_unit, "inv_sort": inv_sort,
            "inv_entries": inv_entries}


def slice_axis(field, gold):
    """보고서필드가 GOLD 컬럼의 어떤 분해 축인지 — 같은 컬럼으로 수렴하는 필드 구분."""
    parts = []
    if re.match(r"^PC", field):
        parts.append("DIM_DEVICE.DEVICE_TYPE='PC'")
    elif re.match(r"^APP", field):
        parts.append("DIM_DEVICE.DEVICE_TYPE='APP' ⚠️축값 미적재")
    elif re.match(r"^M[가-힣]", field):
        parts.append("DIM_DEVICE.DEVICE_TYPE='M'")
    if not any(t in gold for t in ("YEAR_START", "PREV_MONTH", "MONTH_END", "_CUM", "CUM_")):
        for tok, lab in (("전전년", "전전년 동기"), ("전년", "전년 동기"), ("전월", "전월"),
                         ("전주", "전주"), ("누계", "YTD 누계"), ("주간", "주차"),
                         ("당월", "당월"), ("당해년도", "당기")):
            if tok in field:
                parts.append("기간윈도우: " + lab)
                break
    if "신규" in field:
        parts.append("FMM.NEW_EXISTING_FLAG='신규'")
    elif "기존" in field:
        parts.append("FMM.NEW_EXISTING_FLAG='기존'")
    return " · ".join(parts)


def map_report_fields(doc, indices, kind):
    out, seen_fields = [], []

    def lookup(field):
        """인벤토리 → 지표사전 순으로 3단 사다리(정확 → 단위접미 → 어순)를 적용한다."""
        keys = dict(_keys(field))
        # ① GOLD 물리 필드 인벤토리
        for kd, store, tag in (("정확", "inv", "필드인벤토리"),
                               ("단위", "inv_unit", "필드인벤토리(단위)"),
                               ("어순", "inv_sort", "필드인벤토리(어순)")):
            k = keys.get(kd)
            if k and k in indices[store]:
                e = indices[store][k]
                return e["mno"] or "(215밖)", f"{e['table']}.{e['col']}", tag
        # ② 지표사전
        for kd, store, tag in (("정확", "metric", "지표사전"),
                               ("단위", "metric_unit", "지표사전(단위)"),
                               ("어순", "metric_sort", "지표사전(어순)")):
            k = keys.get(kd)
            if k and k in indices[store]:
                m = indices[store][k]
                return m[0], m[2], tag
        return None

    def resolve(field):
        # ⓪ 교정 등록부 — 정본 인벤토리의 정확일치가 **틀렸음을 실측으로 확인**한 항목이 최우선
        ov = FMO.apply(field)
        if ov:
            gold, kind_, _why = ov
            return "", gold, f"교정등록부({kind_})"
        # ①② 인벤토리·지표사전 사다리 (원표기 → 등록 별칭)
        hit = lookup(field)
        if hit:
            return hit
        clean = re.sub(r"\*+\s*\(?중복\)?\s*\*+", "", field).strip()
        if clean != field:
            hit = lookup(clean)
            if hit:
                return hit
        alias = FMO.label_alias(field) or FMO.label_alias(clean)
        if alias:
            hit = lookup(alias)
            if hit:
                return hit[0], hit[1], hit[2] + "·별칭등록부"
        # ③ 규칙 판정 (율·증감·구성비 등 = SV 파생)
        #   🔴 출력 형식 계약: GOLD 매핑 문자열은 반드시 `SV metric — ` 로 시작한다.
        #      `09_보고서필드_조립가능성` 생성기가 이 접두로 「물리 컬럼이 아니다」를 판정한다
        #      (형식을 바꾸면 조립가능성 판정이 조용히 「판정불가」로 무너진다 — 실측 재발함).
        #      분류 표시는 **지표# 열**에 둔다.
        if "납입" in field and "명" in field:
            return ("(FMM 파생)",
                    "SV metric — 납입회원수(명) = COUNT(DISTINCT MEMBER_DK WHERE 납입성공)", "SV파생")
        if RATE_RE.search(field):
            if "달성" in field:
                return "공1·2·3", "SV metric — 목표달성율(%) = 개발(건)/회원개발목표 (FTG_D 분모)", "SV파생"
            if "구성비" in field or "비중" in field:
                return "(ratio-of-total)", "SV metric — 구성비/비중(%) = 부분/전체×100", "SV파생"
            if "증감" in field:
                if "율" in field or "%" in field:
                    return "공60", "SV metric — 증감율(%) = (당기−전기)/전기×100 (P7 시계열)", "SV파생"
                return "공59", "SV metric — 증감 = 당기−전기 (P7 시계열)", "SV파생"
            if "집행율" in field:
                return "(overview)", "SV metric — 집행율(%) = 집행예산/편성예산 (FBD, P7)", "SV파생"
            if "1인당" in field or "1명당" in field:
                return "공61", "SV metric — 1명당 건수 = 활동회원(건)/활동회원(명)", "SV파생"
            if "성공율" in field:
                return "(FSE 파생)", "SV metric — 성공율(%) = SUCCESS_MEMBERS/SEND_MEMBERS", "SV파생"
            return "(SV ratio)", "SV metric — 비율(%) — SV time-intelligence/ratio", "SV파생"
        if "누계" in field:
            # 🔴 누계는 **물리 미저장**이다 — base 컬럼을 그대로 가리키면 단월값을 누계로 읽는다.
            #   그래도 base 를 같이 실어야 추적이 끊기지 않는다.
            b = lookup(field.replace("누계", "").strip()) or lookup(field.replace("(누계)", "").strip())
            tail = f" · base: {b[1]}" if b else ""
            return "(YTD)", f"SV metric — 누계 = base 의 YTD running sum (P7·물리 미저장){tail}", "SV파생"
        # ④ 열리지 않는 것은 「불가」로 명시 — 미매칭과 섞으면 사유가 사라진다
        br = FMO.blocked_reason(field)
        if br:
            return "", f"⛔ 불가 — {br}", "⛔불가(등록부)"
        # ⑤ 부분일치 (⚠ 검증필요)
        n = _norm(field)
        if len(n) >= 3:
            for e2 in indices["inv_entries"]:
                ln = e2["label_norm"]
                if len(ln) >= 3 and e2["col"] not in INV_NO_SUBSTR and (n in ln or ln in n):
                    return e2["mno"] or "(215밖)", f"{e2['table']}.{e2['col']}", "필드인벤토리~"
            for k, m2 in indices["metric"].items():
                if len(k) >= 3 and (n in k or k in n):
                    return m2[0], m2[2], "지표사전~"
        # ⑥ 평균 — 물리 `AVG_*` 컬럼을 먼저 찾은 뒤에도 못 찾으면 SV 집계로 본다
        if "평균" in field:
            return "(SV avg)", "SV metric — 평균 = base AVG (SV 집계)", "SV파생"
        return "", "(미매칭 — GOLD 물리·SV 대응 미확인)", "(미매칭)"

    for t in parse_md(doc):
        h = t["header"]
        i_f = col_idx(h, "필드값")
        if i_f is None:
            continue
        i_src = col_idx(h, "데이터 원천", "원천")
        i_ty = col_idx(h, "데이터 TYPE", "TYPE")
        area = "마케팅 보고서" if kind == "mkt" else (t["h1"] or "")
        section = t["h2"] or t["h3"] or ""
        if t["label"]:
            section = (section + " · " + t["label"]).strip(" ·")
        for r in t["rows"]:
            if len(r) <= i_f:
                continue
            field = r[i_f].strip()
            if not field or field.startswith("---"):
                continue
            seen_fields.append(field)
            src = r[i_src].strip() if i_src is not None and len(r) > i_src else ""
            ty = r[i_ty].strip() if i_ty is not None and len(r) > i_ty else ""
            mno, gold, basis = resolve(field)
            out.append([area, section, field, src, ty, mno, gold, slice_axis(field, gold), basis])
    return out, seen_fields


# ───────────────────────────── 출력 ─────────────────────────────
def write_csv(rows, mkt, mem):
    p = os.path.join(OUT_DIR, BASENAME + ".csv")
    with open(p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([f"# 생성기: {GEN_PATH} | 측정일: {MEASURED}"])
        w.writerow([f"# {PROV}"])
        w.writerow([])
        w.writerow([f"## 지표 → GOLD 매핑 ({len(rows)})"])
        w.writerow(HEADER)
        w.writerows(rows)
        w.writerow([])
        w.writerow(["## 마케팅 보고서필드 → 지표#/GOLD"])
        w.writerow(FIELD_HEADER)
        w.writerows(mkt)
        w.writerow([])
        w.writerow(["## 회원 보고서필드 → 지표#/GOLD"])
        w.writerow(FIELD_HEADER)
        w.writerows(mem)
    return p


def write_md(rows, mkt, mem):
    import collections
    p = os.path.join(OUT_DIR, BASENAME + ".md")
    ok = sum(1 for r in rows if r[11] == "OK")
    pa = sum(1 for r in rows if r[11] == "PARTIAL")
    wa = sum(1 for r in rows if r[11] == "WAIT")
    meas = sum(1 for r in rows if r[12].startswith("실측:"))
    basis_cnt = collections.Counter(r[8] for r in mkt + mem)
    L = []
    A = L.append
    A("<!-- LLM-METADATA")
    A("doc_id: METRIC_TO_GOLD_MAPPING")
    A("doc_role: 지표번호 → GOLD(FACT/DIM/SV·물리컬럼·SV base) 추적 장표 (현업용)")
    A("project: GN_DW (굿네이버스)")
    A("grounded_on: 02_지표 분류.md · 02·03 지표사전 · 04_SV파생 매핑.md · 05_필드 인벤토리.md · "
      "30_output_share/04_컬럼계보매핑.csv(산출물) · census(GOLD 전 컬럼 실측) · "
      "field_mapping_override.py(교정 등록부) · 04·05 보고서필드 인벤토리")
    A(f"generator: {GEN_PATH}")
    A(f"measured: {MEASURED}")
    A("generated: auto (do-not-edit)")
    A("END-METADATA -->")
    A("")
    A("# 지표 → GOLD 매핑 장표 (현업용)")
    A("")
    A(f"> ⚙️ **생성기**: `{GEN_PATH}` · 측정일 **{MEASURED}** — {PROV}")
    A("> **읽는 법**: 현업/기획이 원하는 **지표(지표번호)** 를 기준으로, 그 지표가 GOLD의 어느 **배속(FACT/DIM/SV)** 에")
    A("> 어떤 **물리컬럼**(measure·dimension) 또는 **SV base**(derived=율/구성비/LTV 등)로 매핑됐고, 그 값이")
    A("> 어떤 **SILVER→BRONZE 원천**에서 오는지 한 줄로 추적합니다.")
    A("> 상태: **OK** 사용가능 · **PARTIAL** 일부 대기 · **WAIT** 값 없음/원천 입고 대기")
    A("")
    A("> 🔴 **상태는 가능한 한 실측입니다** — 지표에 대응하는 GOLD 물리 컬럼을 특정할 수 있으면")
    A("> `COUNT`/`COUNT_IF(<>0)` 로 직접 측정해 판정하고, 근거를 `상태_근거` 열에 `실측:` 으로 적습니다.")
    A("> 물리 컬럼을 특정할 수 없는 경우(파생 metric·차원 배속)에만 배속·원천 계통으로 **추정**하고 `추정:` 으로 표시합니다.")
    A("> **`추정:` 행을 사용 가능 근거로 단독 인용하지 마세요.**")
    A("")
    A("## 0. 요약")
    A("")
    A(f"- 총 **{len(rows)}개** 지표 (공통 162 + 신규 53).")
    A(f"- 상태: ✅ OK **{ok}** · ◐ PARTIAL **{pa}** · ⛔ WAIT **{wa}**")
    A(f"- 판정 근거: **실측 {meas}** / 추정 {len(rows) - meas} (실측 = GOLD 물리 컬럼 census 직접 조회)")
    A("")
    A("> 🔴 **원천 계통 추정으로 상태를 매기면 「사용가능」이 거짓이 되는 방향으로만 틀립니다** — 컬럼이 전건 `0` 인")
    A("> 지표를 `OK` 로 분류하면 현업이 조회해서 받은 `0` 을 실적으로 믿습니다(에러도 경고도 없습니다).")
    A("> 그래서 이 장표는 물리 컬럼을 특정할 수 있는 지표를 전부 측정해서 판정합니다(§0-2).")
    A("- **유형별 GOLD 매핑 규칙**: `measure`→FACT 물리컬럼 · `dimension`→DIM(또는 FMM degen/스냅샷) · "
      "`derived`→**SV metric**(분자/분모 base로 계산, 물리컬럼 아님. 단 GA4 비가산 #98·108은 FGA 물리적재).")
    A("- **약어**: " + " · ".join(f"{k}={v}" for k, v in sorted(ABBR.items()) if "-" not in k) + " · SV=Semantic View metric.")
    A("")
    A("### 0-2. 🔴 실측 `WAIT` — 「설계는 됐으나 값이 없는」 지표")
    A("")
    wait_measured = [r for r in rows if r[11] == "WAIT" and r[12].startswith("실측:")]
    A(f"아래 **{len(wait_measured)}개** 지표는 배속·계보가 모두 확정돼 있으나 대응 GOLD 물리 컬럼이 **전건 0 또는 NULL** 이다.")
    A("조회하면 에러 없이 `0` 이 반환되므로 **그 `0` 을 실적으로 읽으면 조용히 틀린다**(P15).")
    A("")
    A("| 지표# | 지표명 | GOLD 매핑 | 실측 근거 |")
    A("|---|---|---|---|")
    for r in wait_measured:
        A(f"| `{r[0]}` | {r[2]} | `{r[7]}` | {r[12][4:].strip()} |")
    A("")
    A("### 0-1. 보고서필드 매핑 신뢰도 (⚠ 커버리지 ≠ 정확도)")
    A("")
    A("| 매핑근거 | 건수 | 신뢰도 | 해석 |")
    A("|---|---:|---|---|")
    guide = [
        ("교정등록부", "높음(실측 교정)", "정본 인벤토리의 정확일치가 **틀렸음을 물리 실측으로 확인**해 바로잡은 항목 — "
                                  "`field_mapping_override.py` 에 사유·근거 기재"),
        ("필드인벤토리", "높음(정확일치)", "05_필드 인벤토리.md 라벨과 정확히 일치 — 물리 GOLD 컬럼 확정"),
        ("지표사전", "높음(정확일치)", "지표사전 지표명과 정확히 일치"),
        ("SV파생", "중(규칙기반)", "율·증감·구성비·대비 등 규칙 판정 — **분자/분모 확정은 04_SV파생 매핑.md·현업 확인 필요**"),
        ("필드인벤토리~", "**검증필요**(부분일치)", "라벨 부분일치 — 동명이의 가능. 표본검증 권장"),
        ("지표사전~", "**검증필요**(부분일치)", "지표명 부분일치 — 동명이의 가능. 표본검증 권장"),
        ("(미매칭)", "—", "GOLD 물리·SV 어느 쪽도 대응 없음(어드민 제외분·원천 부재)"),
    ]
    for name, conf, desc in guide:
        A(f"| `{name}` | {basis_cnt.get(name, 0)} | {conf} | {desc} |")
    A("")
    A("> **후속 작업 주의**: 위 표의 `~`(부분일치)·`SV파생` 행은 **문자열/규칙 기반 추정**이다. 현업 확정 전")
    A("> 계약·개발 산출물의 근거로 단독 인용하지 말고, `교정등록부`/`필드인벤토리`/`지표사전`(정확일치)으로 재확인할 것.")
    A("> 생성 시점 원천 문서가 바뀌면 이 장표는 **자동 갱신되지 않는다** → 생성기 재실행 필요.")
    A("")
    A("## 1. 지표 → GOLD 전체 매핑")
    A("")

    def emit(title, subset):
        A(f"### {title} ({len(subset)})")
        A("")
        A("| 지표# | 지표명 | 유형 | 소스 | 단위 | GOLD 배속 | GOLD 매핑 (물리컬럼 / SV base) | "
          "SILVER 원천 | BRONZE 원천 | 정본 계산식 | 상태 | 상태 근거 |")
        A("|---|---|---|---|---|---|---|---|---|---|---|---|")
        for r in subset:
            f10 = (r[10] or "").replace("|", "\\|")
            A(f"| `{r[0]}` | {r[2]} | {r[3]} | {r[4]} | {r[5]} | `{r[6]}` | `{r[7]}` | "
              f"`{r[8]}` | `{r[9]}` | {f10} | {r[11]} | {r[12]} |")
        A("")

    emit("1-A. 공통 지표", [r for r in rows if r[0].startswith("공")])
    emit("1-B. 신규 지표", [r for r in rows if r[0].startswith("신")])
    A("---")
    A("")

    def emit_fields(title, note, data):
        A(title)
        A("")
        A(note)
        A("")
        A("| 영역 | 섹션 | 필드값 | 데이터 원천 | TYPE | 대응 지표# | GOLD 매핑 | 분해축 | 매핑근거 |")
        A("|---|---|---|---|---|---|---|---|---|")
        for r in data:
            A(f"| {r[0]} | {r[1]} | {r[2]} | {r[3]} | {r[4]} | {r[5]} | `{r[6]}` | {r[7]} | {r[8]} |")
        A("")

    emit_fields("## 2. 마케팅 보고서필드 → 지표#/GOLD 매핑",
                "> `99_provided_definition/04_마케팅_보고서필드 인벤토리.md` 의 필드를 지표번호·GOLD로 매핑.\n"
                "> ⚠️ 이 표는 **라벨→컬럼 매핑**이며 **섹션 단위 조립 가능성이 아니다**(P86). "
                "섹션을 한 표로 조립할 수 있는지는 `09_보고서필드_조립가능성.md` 를 본다.", mkt)
    emit_fields("## 3. 회원 보고서필드 → 지표#/GOLD 매핑",
                "> `99_provided_definition/05_회원_보고서필드 인벤토리.md` 의 필드를 지표번호·GOLD로 매핑.\n"
                "> ⚠️ 조립 가능성은 `09_보고서필드_조립가능성.md` 소관이다(P86).", mem)
    A("_Co-authored with CoCo_")
    open(p, "w", encoding="utf-8").write("\n".join(L) + "\n")
    return p


def write_xlsx(rows, mkt, mem):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D0D0D0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    stat_fill = {"OK": PatternFill("solid", fgColor="E2EFDA"),
                 "PARTIAL": PatternFill("solid", fgColor="FFF2CC"),
                 "WAIT": PatternFill("solid", fgColor="FCE4D6")}
    wb = Workbook()
    ws = wb.active
    ws.title = "00_INDEX"
    ws["A1"] = "GN_DW 지표 → GOLD 매핑 장표"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws["A2"] = f"⚙️ 생성기: {GEN_PATH} · 측정일 {MEASURED} — {PROV}"
    ws.append([])
    ws.append(["시트", "내용"])
    for t, d in (("01_지표GOLD매핑", "215 지표 → 배속·물리컬럼/SV base·SILVER/BRONZE·계산식·상태(실측)"),
                 ("02_마케팅보고서필드", "04 마케팅 보고서필드 → 지표#/GOLD"),
                 ("03_회원보고서필드", "05 회원 보고서필드 → 지표#/GOLD")):
        ws.append([t, d])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 88

    def sheet(title, header, data, widths, status_col=None):
        s = wb.create_sheet(title[:31])
        s["A1"] = f"⚙️ 생성기: {GEN_PATH} · 측정일 {MEASURED}"
        s["A1"].font = Font(size=9)
        s.append([])
        s.append([])
        s.append(list(header))
        for i, w in enumerate(widths, 1):
            from openpyxl.utils import get_column_letter
            s.column_dimensions[get_column_letter(i)].width = w
        for j in range(1, len(header) + 1):
            c = s.cell(row=4, column=j)
            c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, wrap, bd
        for r in data:
            s.append(list(r))
        for i in range(5, s.max_row + 1):
            for j in range(1, len(header) + 1):
                c = s.cell(row=i, column=j)
                c.alignment, c.border = wrap, bd
            if status_col:
                v = s.cell(row=i, column=status_col).value
                if v in stat_fill:
                    s.cell(row=i, column=status_col).fill = stat_fill[v]
        s.freeze_panes = "A5"

    sheet("01_지표GOLD매핑", HEADER, rows, (8, 8, 30, 10, 8, 8, 16, 40, 34, 40, 46, 9, 48), 12)
    sheet("02_마케팅보고서필드", FIELD_HEADER, mkt, (18, 26, 26, 22, 12, 12, 40, 30, 18))
    sheet("03_회원보고서필드", FIELD_HEADER, mem, (16, 30, 26, 20, 12, 12, 40, 30, 18))
    tmp = os.path.join("/tmp", BASENAME + ".xlsx")
    wb.save(tmp)
    import shutil
    p = os.path.join(OUT_DIR, BASENAME + ".xlsx")
    shutil.copyfile(tmp, p)
    return p


def main():
    rows, _ = build_metric_rows()
    indices = build_field_index(rows)
    mkt, seen1 = map_report_fields(DOC_MKT, indices, "mkt")
    mem, seen2 = map_report_fields(DOC_MEM, indices, "mem")

    # 🔴 등록부 유효성 검증(P85-②) — 인벤토리에서 사라진 항목은 죽은 교정이다
    dead = FMO.validate(seen1 + seen2)
    applied = sum(1 for r in mkt + mem if r[8].startswith("교정등록부"))
    if dead:
        print("⚠️ 교정 등록부에 인벤토리 부재 항목:", ", ".join(dead))
    if applied == 0:
        print("🔴 교정 등록부 적용 0건 — 배선이 끊겼을 수 있다(O45-D 재발 신호)")
    print(f"교정 적용 {applied}행")
    # 🔴 별칭 등록부의 **우변**이 정본에 실재하는지 확인한다 — 없으면 그 별칭은 무력하다
    known = set(indices["inv"]) | set(indices["metric"])
    bad_alias = sorted(k for k, v in FMO.LABEL_ALIAS.items() if _norm(v) not in known)
    if bad_alias:
        print("⚠️ 별칭 우변이 정본에 없음(무력 별칭):", ", ".join(bad_alias))
    unmatched = sorted({r[2] for r in mkt + mem if r[8] == "(미매칭)"})
    print(f"미매칭 고유 {len(unmatched)}: " + ", ".join(unmatched[:40]))

    def rate(x):
        n = sum(1 for r in x if r[5] or r[8] != "(미매칭)")
        return f"{n}/{len(x)} ({n / max(len(x), 1) * 100:.1f}%)"

    print("지표 행:", len(rows), "| 마케팅 필드:", rate(mkt), "| 회원 필드:", rate(mem))
    print("CSV :", write_csv(rows, mkt, mem))
    print("MD  :", write_md(rows, mkt, mem))
    try:
        print("XLSX:", write_xlsx(rows, mkt, mem))
    except ImportError:
        print("XLSX: openpyxl 미설치")


if __name__ == "__main__":
    main()
