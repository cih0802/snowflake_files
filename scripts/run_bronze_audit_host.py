# BRONZE 노출감사 호스트 실행 러너 — Snowflake 직접조회 또는 캐시 폴백으로 감사 산출물 생성
# Co-authored with CoCo
"""
gen_bronze_exposure_audit.py 의 판정 로직을 호스트 환경에서 실행하는 러너.

실행 모드:
  인수 없이 → Snowflake 직접조회 (snowflake-connector-python + OAuth 토큰)
  인수 있으면 → 기존 캐시 폴백 (호환용)

출력: 30_output_share/06_BRONZE노출감사.{md,csv,xlsx}

dbt 하드코딩 검출은 workspace 파일을 직접 스캔한다(호스트는 read 가능).
"""
import csv, os, re, sys, glob
from datetime import date

# 기본은 workspace FUSE 마운트. 마운트 장애 시 스테이지에서 내려받은 사본 경로를
# GN_DW_WS 로 지정해 동일 로직을 그대로 재현한다(입력 경로만 다르고 판정은 불변).
WS = os.environ.get("GN_DW_WS", "/workspace")
OUT_DIR = os.environ.get("GN_DW_OUT", os.path.join(WS, "30_output_share"))
BASENAME = "06_BRONZE노출감사"
GEN_PATH = "scripts/gen_bronze_exposure_audit.py"
RUNNER_PATH = "scripts/run_bronze_audit_host.py"
PROV = f"자동 생성물 — 생성기 {GEN_PATH}(+러너 {RUNNER_PATH}) 재실행으로 갱신. 직접 편집 금지."
AUDIT_DATE = date.today().isoformat()

CACHE_DIR = sys.argv[1] if len(sys.argv) > 1 else None

EXCLUDE_PATTERNS = [
    r"^(FRST_REGISTER|LST_RGSTR|FRST_REGIST|LST_UPDT|REG_DATE|UPD_DATE|DEL_YN|USE_YN).*",
    r"^(RGSTR_ID|UPDT_ID|REGISTER_ID|UPDATER_ID|FILLER.*)$",
    r"^(SND_MSG|DISCHARGE_REPORT|LETTER_CONTENT).*",
    r"^(MBER_NM|CHILD_NM_KOR|CHILD_NM_ENG|ADDR|ZIP|TEL|HP|MOBILE|FAX|EMAIL_ADDR).*",
]
EXCLUDE_RE = re.compile("|".join(EXCLUDE_PATTERNS), re.IGNORECASE)
DW_META_COLS = {"DW_SOURCE_SYSTEM", "DW_SOURCE_TABLE", "DW_LOAD_TS", "DW_UPDATE_TS", "DW_BATCH_ID"}

# ── 계보 매핑 ((BRONZE 테이블, BRONZE 컬럼) → (GOLD 컬럼, GOLD 모델파일)) ──
# P13 대응: 개명 적재는 이름매칭으로 탐지 불가하므로 명시 등록한다.
# ⚠️ 키에 BRONZE 테이블을 포함하는 이유: 컬럼명만으로는 원천 간 충돌이 발생한다
#    (예: AGENCY.DEVICE 는 FACT_AD_PERFORMANCE 계보로 하드코딩 상태이나,
#     GA4.device 는 FACT_GA_BEHAVIOR 계보로 정상 적재 — 동일시하면 오판).
# ⚠️ 값에 GOLD 모델을 포함하는 이유: 동명 GOLD 컬럼이 여러 모델에 있고 한쪽만 하드코딩인 경우가 있다
#    (예: AD_COST — FACT_AD_PERFORMANCE 실적재 / FACT_BUDGET 하드코딩).
# 테이블명은 접두 매칭 (GA4 events_YYYYMMDD 샤드 대응).
# 근거: 2026-07-27~28 BRONZE 실측 + dbt 모델 대조.
LINEAGE_MAP = {
    # ── AGENCY VIDEO — 방송(영상)광고 속성 → FACT_AD_BROADCAST / FACT_AD_PERFORMANCE ──
    ("VIDEO_AD_CMPGN_DTLS", "TIME_RNG"):           ("TIME_BAND",           "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "CM_AREA"):            ("CM_POSITION",         "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "AD_STRT_TIME"):       ("AD_START_TIME",       "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "BRDC_DATE"):          ("BROADCAST_DATE",      "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "AD_SEC"):             ("DURATION_SEC",        "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "SCHDL_NM"):           ("PROGRAM_NM",          "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "CHNNL_NM"):           ("CHANNEL_COMPANY",     "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "CHNNL_CMPNY_TY_NM"): ("CHANNEL_COMPANY_TYPE", "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "SPOT_TY"):            ("SPOT_TYPE",           "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "DAY_DIV_NM"):          ("DAY_DIV",             "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "PRG_STRT_TIME"):      ("PRG_START_TIME",      "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "CTV_DIV_NM"):          ("CTV_DIV",             "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "CONV_CALL_CNT"):      ("CONV_CALL_CNT",       "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "AD_VIEW_RT"):         ("AD_VIEW_RT_SRC",      "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "CPC"):                ("CPC_SRC",             "FACT_AD_BROADCAST.sql"),
    ("VIDEO_AD_CMPGN_DTLS", "ACTL_PUR_AD_COST_KRW"): ("AD_COST",           "FACT_AD_PERFORMANCE.sql"),
    # ── AGENCY REBRDC — 재방송광고 속성 / 사례 / 개발실적 ──
    ("REBRDC_AD_CMPGN_DTLS", "TIME_RNG_DIV_NM"):   ("TIME_BAND",           "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "BRDC_TIME"):         ("TIME_BAND",           "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "RE_BRDC_TY_NM"):     ("RT_TYPE",             "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "DATE"):              ("BROADCAST_DATE",      "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "BRDC_NM"):           ("PROGRAM_NM",          "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CHNNL_CMPNY"):       ("CHANNEL_COMPANY",     "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "BRDC_DIV_NM"):        ("BRDC_DIV",            "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "DVLP_MBER_CNT"):     ("DVLP_MEMBER_CNT",     "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "DVLP_CNT"):          ("DVLP_CNT",            "FACT_AD_BROADCAST.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE1_BSNS_DIV_NM"):      ("BIZ_DIV",       "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE1_FAM_TY_NM"):        ("FAMILY_TYPE",   "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE1_APPEAL_POINT_NM"):  ("APPEAL_POINT",  "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE1_CASE_DIV_NM"):      ("CASE_DIV",      "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE2_BSNS_DIV_NM"):      ("BIZ_DIV",       "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE2_FAM_TY_NM"):        ("FAMILY_TYPE",   "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE2_APPEAL_POINT_NM"):  ("APPEAL_POINT",  "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE2_CASE_DIV_NM"):      ("CASE_DIV",      "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE3_BSNS_DIV_NM"):      ("BIZ_DIV",       "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE3_FAM_TY_NM"):        ("FAMILY_TYPE",   "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE3_APPEAL_POINT_NM"):  ("APPEAL_POINT",  "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "CASE3_CASE_DIV_NM"):      ("CASE_DIV",      "FACT_AD_BROADCAST_CASE.sql"),
    ("REBRDC_AD_CMPGN_DTLS", "BRDC_SCHDL_COST"):        ("AD_COST",       "FACT_AD_PERFORMANCE.sql"),
    # ── AGENCY DGT — 디지털광고 기기 / 코어 측정지표 / 위성 속성 / 파생원본 ──
    ("DGT_AD_CMPGN_DTLS", "DEVICE"):               ("DEVICE_SK",           "FACT_AD_PERFORMANCE.sql"),
    ("DGT_AD_CMPGN_DTLS", "EXPS_CNT"):             ("IMPRESSIONS",         "FACT_AD_PERFORMANCE.sql"),
    ("DGT_AD_CMPGN_DTLS", "CLICK_CNT"):            ("CLICKS",              "FACT_AD_PERFORMANCE.sql"),
    ("DGT_AD_CMPGN_DTLS", "GA_CONV_MBER_CNT"):     ("GA_CONV_MEMBERS",     "FACT_AD_PERFORMANCE.sql"),
    ("DGT_AD_CMPGN_DTLS", "CONV_VU_CNT"):          ("GA_CONV_CNT",         "FACT_AD_PERFORMANCE.sql"),
    ("DGT_AD_CMPGN_DTLS", "GA_AD_COST"):           ("AD_COST",             "FACT_AD_PERFORMANCE.sql"),
    ("DGT_AD_CMPGN_DTLS", "READ_CNT"):             ("READ_CNT",            "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "MEDIA_PTNT_CUST_CNT"):  ("MEDIA_POTENTIAL_CUST_CNT", "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "CRM_DVLP_CNT"):         ("CRM_DEV_CNT",         "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "PAGE_TYPE_NM"):          ("PAGE_TYPE",           "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "AD_GRP_NM"):            ("AD_GROUP_NM",         "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "GRP_DIV_NM"):           ("GROUP_DIV",           "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "MATR_TY_NM"):           ("CREATIVE_TYPE",       "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "AD_TY_NM"):             ("AD_TYPE_NM",          "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "CTR"):                  ("CTR_SRC",             "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "CVR"):                  ("CVR_SRC",             "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "CPC"):                  ("CPC_SRC",             "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "CPM"):                  ("CPM_SRC",             "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "CPA"):                  ("CPA_SRC",             "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "DEV_UNIT_PRICE"):        ("DEV_UNIT_PRICE_SRC",  "FACT_AD_DIGITAL.sql"),
    ("DGT_AD_CMPGN_DTLS", "VTR"):                  ("VTR_SRC",             "FACT_AD_DIGITAL.sql"),
    # ── GA4 — VARIANT 원천이 GA4_DEVICE→DIM_DEVICE→FACT_GA_BEHAVIOR 로 정상 적재 ──
    ("events_", "device"):    ("DEVICE_SK",    "FACT_GA_BEHAVIOR.sql"),
    ("events_", "platform"):  ("DEVICE_SK",    "FACT_GA_BEHAVIOR.sql"),
    ("events_", "traffic_source"): ("GA_SOURCE_SK", "FACT_GA_BEHAVIOR.sql"),
    ("events_", "event_name"):     ("GA_EVENT_SK",  "FACT_GA_BEHAVIOR.sql"),
    # ── CRM 개명 계보 (2026-08-06 등재 · O38·O45 배선분) ──
    # 🔴 P90 ②: 신규 객체를 만들면 이 등록부가 그 사실을 모른다. 아래 8건은 등재 전까지
    #    전부 「SILVER까지만 · GOLD 미승격」으로 오판됐다(2026-08-06 실행에서 실측 확인).
    #    근거는 dbt 모델 본문 대조이며 각 행에 GOLD 컬럼·모델을 명시한다.
    # O45 마케팅캠페인 conformed 축 — 광고(AGENCY) ↔ 개발실적(CRM) 결합축
    ("TM_CM_MKTNG_CMPGN_MNG", "MK_CMPGN_CD"): ("MKTG_CAMPAIGN_BK",   "DIM_MARKETING_CAMPAIGN.sql"),
    ("TM_CM_MKTNG_CMPGN_MNG", "MK_CMPGN_NM"): ("MKTG_CAMPAIGN_NAME", "DIM_MARKETING_CAMPAIGN.sql"),
    # 캠페인 마스터측 조인키(NUMBER) → 브리지 조인으로 SK 승격 (DIM_CAMPAIGN.sql:65)
    ("TM_CM_CMPGN_MNG", "MKTG_CMPGN_NM"):     ("MKTG_CAMPAIGN_SK",   "DIM_CAMPAIGN.sql"),
    # O38 실적부서 · O45 후원사업 배선 (개발/중단 사건 팩트)
    ("TM_MM_FDRM_MBER_DVLP_AMT", "ACMSLT_DEPT_CD"): ("ORG_SK",          "FACT_MEMBER_EVENT.sql"),
    ("TM_MM_FDRM_MBER_DVLP_AMT", "SPNSR_BSNS_ID"):  ("SPONSORSHIP_SK",  "FACT_MEMBER_EVENT.sql"),
    # O45 회비 팩트 — 회비·기부금 UNION 양쪽에서 후원사업축이 올라온다(CRM_PAYMENT_BILLING)
    ("TM_PM_MBRFEE_ACMSLT", "SPNSR_BSNS_ID"): ("SPONSORSHIP_SK", "FACT_MEMBER_FEE.sql"),
    ("TM_PM_MBRFEE_ACMSLT", "MBRFEE_DIV_CD"): ("FEE_DIV_CD",     "FACT_MEMBER_FEE.sql"),
    ("TM_PM_DNTN_DTLS",     "SPNSR_BSNS_ID"): ("SPONSORSHIP_SK", "FACT_MEMBER_FEE.sql"),
    # 후원사업 마스터 → 차원 자연키
    ("TM_CM_SPNSR_BSNS_INFO", "SPNSR_BSNS_ID"): ("SPONSORSHIP_BK", "DIM_SPONSORSHIP.sql"),
    # ── CRM 개명 계보 (2026-08-13 O65 등재 · O59-N DEC-35 2단계 축B 배선분) ──
    # 🔴 왜 등재하는가: 이 2건은 `classify()` 의 (5) 동명 매칭에 걸리지 않고 `silver_refs`(SILVER SQL
    #    토큰) 경로로 떨어져 「SILVER까지만 · 중간(SQL참조)」이 됐다. 그러나 **GOLD 도달이 실측 확증**된다:
    #      · SILVER 에 동명 컬럼 없음 — INFORMATION_SCHEMA 조회 0행(이름으로는 찾을 수 없다)
    #      · `CRM_SEND_MEMBER.sql` 28·41행이 두 컬럼을 **축B `SEND_RESULT_CD` 로 개명 적재**
    #      · `FACT_SERVICE_EVENT.sql:55` 가 그것을 투영 · 라이브 채움을 채널로 분해하면 귀속이 갈린다:
    #        MSG_AT(`TRNSMS_FAILR_CD_ID`) **18,439,718** · SND(`CALL_STATUS`) **7,815,657**
    #        · EMAIL·PSTMTR 은 모델이 NULL 고정이라 0 (합 26,255,375 / 전체 38,470,780)
    #    ⇒ P90 ② 의 재현이다(신규 배선을 만들면 이 등록부가 그 사실을 모른다). 2026-08-06 8건과 같은 유형.
    # ⚠️ 이 2건은 「실측으로 확정된 것」만이다. 같은 버킷(SILVER까지만 · 신뢰도 「높음」 아님)에
    #    95건이 남아 있고 그것은 **후보**이지 오류 확정이 아니다 — 전수 판정은 별건(원장 O65 잔여).
    ("TD_MS_MSG_AT_SNDNG_DTLS", "TRNSMS_FAILR_CD_ID"): ("SEND_RESULT_CD", "FACT_SERVICE_EVENT.sql"),
    ("SND_MEMBER_LIST",         "CALL_STATUS"):        ("SEND_RESULT_CD", "FACT_SERVICE_EVENT.sql"),
}


def lookup_lineage(table_name, col_upper):
    """테이블 스코프 계보 조회 (테이블명 접두 매칭 — GA4 샤드 대응)."""
    for (tbl_pat, c), v in LINEAGE_MAP.items():
        if c.upper() != col_upper:
            continue
        if table_name == tbl_pat or table_name.startswith(tbl_pat):
            return v
    return None

# ── 파생 대체 (BRONZE 원본 컬럼 대신 다른 컬럼의 파생값을 적재) ──
# SILVER AGENCY_AD_PERFORMANCE 주석: "연·월 = DATE 파생(텍스트 파싱 금지)".
# BRONZE 텍스트 YEAR/MONTH 등은 의도적으로 미사용 — '미노출'이 아니라 '대체노출'로 판정해야 정확하다.
DERIVED_REPLACED = {
    "YEAR":     "DATE 파생(YEAR(AD_DATE)) 로 대체 — 텍스트 파싱 금지 원칙",
    "MONTH":    "DATE 파생(MONTH(AD_DATE)) 로 대체",
    "BRDC_MT":  "DATE 파생(MONTH) 로 대체",
    "WEEK":     "DATE 파생(WEEKOFYEAR) 로 대체",
    "DAY":      "DATE 파생(DAY) 로 대체",
    "DOW":      "DATE 파생(DAYNAME) 로 대체",
}

# ── 동명이의 위험 컬럼명 ──
# 층 간 단순 이름매칭 시 무관한 GOLD 컬럼과 충돌하는 일반명사.
# 이 이름들은 LINEAGE_MAP/DERIVED_REPLACED 로 명시 등록되지 않은 한 '노출됨' 판정을 금지한다.
AMBIGUOUS_GENERIC = {
    "YEAR", "MONTH", "DAY", "WEEK", "DATE", "TIME", "DOW", "QUARTER",
    "CM", "AGE", "GENDER", "REGION", "TEAM", "CORP", "DIVISION", "DEPARTMENT",
    "PLATFORM", "CHANNEL", "SUBTYPE", "CREATIVE", "BRAND", "DIV_NM",
    "MEMBER_TYPE", "MEMBER_STATUS", "EVENT_NAME", "EVENT_CATEGORY",
    "SPNSR_BSNS_NM", "CMPGN_NM", "MEDIA_NM",
}


def scan_dbt_hardcoding():
    """dbt GOLD 모델 SQL에서 '0 as X_SK' / 'CAST(NULL AS ..) as X' 하드코딩 검출."""
    patterns = [
        re.compile(r"\b0\s+as\s+(\w+_SK)\b", re.IGNORECASE),
        re.compile(r"CAST\s*\(\s*NULL\s+AS\s+\w+(?:\([^)]*\))?\s*\)\s+as\s+(\w+)", re.IGNORECASE),
    ]
    # 결함 2+3 교정: 행 내 모든 alias 를 수집한 뒤 SQL 타입 키워드를 제외한다.
    #   · 결함2(미탐): 이전 정규식은 행말 앵커($)라 한 행에 alias 가 여럿이면 마지막만 검출.
    #   · 결함3(오탐): 앵커를 풀어 모든 `as X` 를 잡으면 `CAST(NULL AS VARCHAR)` 의
    #     타입 키워드를 alias 로 오인. → 타입 키워드 집합으로 걸러낸다.
    alias_pat = re.compile(r"\bas\s+([A-Za-z_]\w*)", re.IGNORECASE)
    SQL_TYPES = {
        "VARCHAR", "CHAR", "CHARACTER", "STRING", "TEXT", "BINARY", "VARBINARY",
        "NUMBER", "DECIMAL", "NUMERIC", "INT", "INTEGER", "BIGINT", "SMALLINT", "TINYINT",
        "FLOAT", "FLOAT4", "FLOAT8", "DOUBLE", "REAL",
        "BOOLEAN", "DATE", "DATETIME", "TIME", "TIMESTAMP", "TIMESTAMP_LTZ",
        "TIMESTAMP_NTZ", "TIMESTAMP_TZ", "VARIANT", "OBJECT", "ARRAY",
    }

    hardcoded = {}   # (model, col) -> info
    real = set()     # (model, col)
    gold_dir = os.path.join(WS, "10_dbt_pipeline", "models", "gold")
    for sql_path in sorted(glob.glob(os.path.join(gold_dir, "**", "*.sql"), recursive=True)):
        if "/target/" in sql_path:
            continue
        model = os.path.basename(sql_path)
        rel = os.path.relpath(sql_path, WS)
        with open(sql_path, encoding="utf-8") as f:
            for lineno, line in enumerate(f, 1):
                hit_cols = set()
                for pat in patterns:
                    for m in pat.finditer(line):
                        col = m.group(1).upper()
                        hit_cols.add(col)
                        if (model, col) not in hardcoded:
                            hardcoded[(model, col)] = {
                                "file": rel, "model": model,
                                "pattern": " ".join(m.group(0).split()), "line": lineno}
                # 결함 2+3 교정: 행 내 모든 alias 수집, SQL 타입 키워드 제외
                for m in alias_pat.finditer(line.split("--")[0]):
                    col = m.group(1).upper()
                    if col not in SQL_TYPES and col not in hit_cols:
                        real.add((model, col))
    return hardcoded, real


def scan_silver_refs():
    """SILVER dbt SQL 토큰 집합 — BRONZE 컬럼이 SILVER SQL에서 참조되는지 판별용."""
    tokens = set()
    silver_dir = os.path.join(WS, "10_dbt_pipeline", "models", "silver")
    for sql_path in glob.glob(os.path.join(silver_dir, "**", "*.sql"), recursive=True):
        if "/target/" in sql_path:
            continue
        with open(sql_path, encoding="utf-8") as f:
            tokens.update(re.findall(r"\b([A-Za-z_][A-Za-z0-9_]{2,})\b", f.read().upper()))
    return tokens


def scan_gold_refs():
    """GOLD dbt SQL 토큰 집합."""
    tokens = set()
    gold_dir = os.path.join(WS, "10_dbt_pipeline", "models", "gold")
    for sql_path in glob.glob(os.path.join(gold_dir, "**", "*.sql"), recursive=True):
        if "/target/" in sql_path:
            continue
        with open(sql_path, encoding="utf-8") as f:
            tokens.update(re.findall(r"\b([A-Za-z_][A-Za-z0-9_]{2,})\b", f.read().upper()))
    return tokens


def parse_cache(path):
    """snowflake_sql_execute 캐시 파일 파싱 → (schema, table, col, dtype, pos) 리스트."""
    rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or line.startswith("TABLE_SCHEMA") \
               or "row(s) returned" in line or line.startswith("["):
                continue
            parts = line.split(",")
            if len(parts) < 5:
                continue
            schema, table, col, dtype = parts[0], parts[1], parts[2], parts[3]
            try:
                pos = int(parts[4])
            except ValueError:
                continue
            rows.append((schema, table, col, dtype, pos))
    return rows


def fetch_columns_snowflake():
    """Snowflake INFORMATION_SCHEMA 직접조회 → (schema, table, col, dtype, pos) 리스트.

    재현성 근거: 세션 캐시(snowflake_sql_execute 출력)는 휘발성이라 동일 결과를
    재생산할 수 없다. 러너는 인수 없이 실행하면 항상 원본 카탈로그를 직접 읽는다.
    """
    import snowflake.connector
    token_path = os.environ.get("SNOWFLAKE_TOKEN_FILE_PATH", "/snowflake/session/token")
    with open(token_path) as f:
        token = f.read().strip()
    conn = snowflake.connector.connect(
        account=os.environ.get("SNOWFLAKE_ACCOUNT", ""),
        host=os.environ.get("SNOWFLAKE_HOST", ""),
        user=os.environ.get("SNOWFLAKE_USERNAME", ""),
        authenticator="oauth",
        token=token,
        role=os.environ.get("SNOWFLAKE_ROLE", "ACCOUNTADMIN"),
        # INFORMATION_SCHEMA 조회는 실행 웨어하우스를 요구한다.
        warehouse=os.environ.get("SNOWFLAKE_WAREHOUSE", "COMPUTE_WH"),
        database="GN_DW",
        insecure_mode=True,  # 샌드박스가 OCSP 응답기를 차단
    )
    sql = """
    SELECT TABLE_SCHEMA, TABLE_NAME, COLUMN_NAME, DATA_TYPE, ORDINAL_POSITION
    FROM GN_DW.INFORMATION_SCHEMA.COLUMNS
    WHERE TABLE_SCHEMA IN ('BRONZE_CRM','BRONZE_AGENCY','BRONZE_ERP','BRONZE_BIGQUERY',
                           'SILVER','GOLD')
    ORDER BY TABLE_SCHEMA, TABLE_NAME, ORDINAL_POSITION
    """
    cur = conn.cursor()
    try:
        cur.execute(sql)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
    return [(r[0], r[1], r[2], r[3], r[4]) for r in rows]


def classify(table_name, col_upper, silver_cols, gold_cols, silver_refs, gold_refs, hardcoded, real):
    """BRONZE 단일 컬럼 노출 판정. 반환 (판정, 신뢰도, 비고).

    판정 우선순위 (아키텍처 정확성 순서):
      1) 제외 (PII·본문·DW메타)
      2) 명시 계보(LINEAGE_MAP) — 모델까지 특정하여 하드코딩/실적재 구분
      3) 파생 대체(DERIVED_REPLACED)
      4) 동명 하드코딩 (전 모델 하드코딩인 경우만)
      5) 동명 GOLD/SILVER 존재 — 단 동명이의 위험명은 판정보류
    """
    if EXCLUDE_RE.match(col_upper):
        return "제외(PII·본문·메타)", "—", "패턴 매칭 제외(감사 범위 외)"
    if col_upper in DW_META_COLS:
        return "제외(DW메타)", "—", "DW 감사 메타컬럼"

    # (2) 명시 계보 — 모델 스코프로 하드코딩 여부 판정
    lin = lookup_lineage(table_name, col_upper)
    if lin:
        gcol, gmodel = lin
        hc = hardcoded.get((gmodel, gcol))
        if hc:
            # 🔴 [2026-08-06] 브랜치별 상이를 「값미주입」으로 뭉개면 안 된다.
            #   같은 모델이 UNION 브랜치별로 실조인과 센티넬을 함께 갖는 경우가 있다
            #   (예: FACT_MEMBER_EVENT — DEV 브랜치는 ORG_SK·SPONSORSHIP_SK 실조인,
            #    STOP 브랜치는 역할 불일치로 0 유지 = O38-B 결정 대기).
            #   이때 「값미주입」이라 답하면 이미 산출 가능한 축을 불가로 안내한다(P61·P76).
            if (gmodel, gcol) in real:
                return ("노출됨(GOLD)", "중간(브랜치별 상이)",
                        f"개명 적재 → GOLD `{gcol}` ({gmodel}) · "
                        f"⚠️ 일부 브랜치는 센티넬 — {gmodel}:{hc['line']} `{hc['pattern']}`")
            return ("⚠️설계O·값미주입", "높음",
                    f"GOLD `{gcol}`({gmodel}) 자리 존재하나 하드코딩 — "
                    f"{gmodel}:{hc['line']} `{hc['pattern']}`")
        if (gmodel, gcol) in real or gcol in gold_cols:
            return "노출됨(GOLD)", "높음", f"개명 적재 → GOLD `{gcol}` ({gmodel})"
        return "미노출(검토대상)", "중간(계보등록)", f"GOLD `{gcol}` 미확인 — 배선 필요"

    # (3) 파생 대체
    if col_upper in DERIVED_REPLACED:
        return "대체노출(파생)", "높음", DERIVED_REPLACED[col_upper]

    # (4) 동명 하드코딩 — 해당 컬럼명이 하드코딩된 모델은 있고 실적재 모델은 없을 때만
    hc_models = [k[0] for k in hardcoded if k[1] == col_upper]
    real_models = [m for (m, c) in real if c == col_upper]
    if hc_models and not real_models:
        hc = hardcoded[(hc_models[0], col_upper)]
        return ("⚠️설계O·값미주입", "높음",
                f"{hc['model']}:{hc['line']} `{hc['pattern']}`")
    if hc_models and real_models:
        return ("노출됨(GOLD)", "중간(모델별 상이)",
                f"실적재 {','.join(sorted(set(real_models)))} / 하드코딩 {','.join(sorted(set(hc_models)))}")

    # (5) 동명 매칭 — 동명이의 위험명은 근거 불충분으로 판정보류
    if col_upper in AMBIGUOUS_GENERIC:
        if col_upper in gold_cols or col_upper in silver_cols:
            return ("판정보류(동명이의)", "낮음(일반명 충돌)",
                    "동명 GOLD/SILVER 컬럼이 있으나 계보 무관 가능 — 실측 필요(P14)")
        return "미노출(검토대상)", "낮음(일반명)", "일반명 — 개별 실측 필요"

    # 결함4 교정(P15): 동명 GOLD 컬럼 존재만으로 '높음'을 주면
    # "설계완료 ≠ 값존재"(P15)와 자기모순. projection 확인 여부로 신뢰도를 분리한다.
    if col_upper in gold_cols:
        if any(c == col_upper for (_m, c) in real):
            return "노출됨(GOLD)", "높음", "동명 GOLD 컬럼 + 실적재 projection 확인"
        return ("노출됨(GOLD)", "중간(스키마만)",
                "GOLD 컬럼 존재하나 dbt projection 미확인 — 값 유무 실측 필요(P15)")
    if col_upper in silver_cols:
        return "SILVER까지만", "높음", "GOLD 미승격"
    if col_upper in silver_refs:
        return "SILVER까지만", "중간(SQL참조)", "SILVER SQL 토큰 참조"
    if col_upper in gold_refs:
        return "SILVER까지만", "낮음(토큰만)", "GOLD SQL 토큰 출현 — 적재 미확인"
    return "미노출(검토대상)", "낮음(이름기반·P13)", "개명·VARIANT param 승격 가능성 — 확정 아님"


VERDICT_ORDER = ["노출됨(GOLD)", "대체노출(파생)", "⚠️설계O·값미주입", "SILVER까지만",
                 "판정보류(동명이의)", "미노출(검토대상)", "제외(PII·본문·메타)", "제외(DW메타)"]


def main():
    print(f"[{AUDIT_DATE}] BRONZE 노출감사 (호스트 실행)")

    # 1) 컬럼 목록 로드 — 결함1 교정: 인수 없으면 Snowflake 직접조회(정본),
    #    인수가 있으면 캐시 폴백(커넥터·토큰 없는 환경 전용).
    all_rows = []
    if not CACHE_DIR:
        print("  모드: Snowflake 직접조회 (OAuth)")
        all_rows = fetch_columns_snowflake()
        print(f"  조회 완료: {len(all_rows)}행")
    else:
        print(f"  모드: 캐시 폴백 ({CACHE_DIR}) — 재현성 없음, 정상 경로 아님")
        for cf in sorted(glob.glob(os.path.join(CACHE_DIR, "snowflake_sql_execute_*.txt"))):
            parsed = parse_cache(cf)
            print(f"  {os.path.basename(cf)}: {len(parsed)}행")
            all_rows.extend(parsed)

    bronze_map, silver_cols, gold_cols = {}, set(), set()
    for schema, table, col, dtype, pos in all_rows:
        if schema.startswith("BRONZE_"):
            bronze_map.setdefault(f"{schema}.{table}", []).append((col, dtype, pos))
        elif schema == "SILVER":
            silver_cols.add(col.upper())
        elif schema == "GOLD":
            gold_cols.add(col.upper())
    bronze_total = sum(len(v) for v in bronze_map.values())
    print(f"  BRONZE {bronze_total}컬럼 · SILVER {len(silver_cols)} · GOLD {len(gold_cols)}")

    # 2) dbt 스캔
    hardcoded, real = scan_dbt_hardcoding()
    silver_refs = scan_silver_refs()
    gold_refs = scan_gold_refs()
    print(f"  dbt 하드코딩 {len(hardcoded)}건 · 실적재 projection {len(real)}건 검출")

    # 3) 판정
    audit_rows, stats = [], {}
    for bronze_key in sorted(bronze_map.keys()):
        source_sys = bronze_key.split(".")[0].replace("BRONZE_", "")
        table_name = bronze_key.split(".")[1]
        for col, dtype, pos in bronze_map[bronze_key]:
            verdict, conf, note = classify(table_name, col.upper(), silver_cols, gold_cols,
                                           silver_refs, gold_refs, hardcoded, real)
            audit_rows.append({
                "원천시스템": source_sys, "BRONZE_테이블": table_name,
                "BRONZE_컬럼": col, "데이터타입": dtype,
                "판정": verdict, "신뢰도": conf, "비고": note,
            })
            stats[verdict] = stats.get(verdict, 0) + 1

    print(f"  판정 완료 {len(audit_rows)}건")
    for v in VERDICT_ORDER:
        if stats.get(v):
            print(f"    {v}: {stats[v]}")

    # 4) 출력
    os.makedirs(OUT_DIR, exist_ok=True)
    write_csv(audit_rows)
    write_md(audit_rows, stats, hardcoded, real)
    write_xlsx(audit_rows, stats, hardcoded, real)
    print(f"  출력 → {OUT_DIR}/{BASENAME}.{{md,csv,xlsx}}")


def write_xlsx(audit_rows, stats, hardcoded, real):
    """현업 배포용 xlsx — 요약 / 하드코딩(최우선) / 원천별 시트."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=13, color="1F4E78")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "노출됨(GOLD)":        PatternFill("solid", fgColor="E2EFDA"),
        "대체노출(파생)":      PatternFill("solid", fgColor="D9E1F2"),
        "⚠️설계O·값미주입":    PatternFill("solid", fgColor="F4B084"),
        "SILVER까지만":        PatternFill("solid", fgColor="FFF2CC"),
        "판정보류(동명이의)":  PatternFill("solid", fgColor="E4DFEC"),
        "미노출(검토대상)":    PatternFill("solid", fgColor="FCE4D6"),
        "제외(PII·본문·메타)": PatternFill("solid", fgColor="D9D9D9"),
        "제외(DW메타)":        PatternFill("solid", fgColor="D9D9D9"),
    }

    def style_hdr(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, wrap, border

    total = len(audit_rows)

    # 00_요약
    ws = wb.active
    ws.title = "00_요약"
    ws["A1"] = f"BRONZE 노출감사 — 전 원천 {total}컬럼 ({AUDIT_DATE})"
    ws["A1"].font = title_font
    ws["A2"] = f"⚙️ {PROV}"
    ws["A3"] = "P13: 이름매칭 기반 — '미노출'은 부재 확정 아님 / P16: 일반명은 판정보류로 격리"
    ws.append([])
    ws.append(["판정", "건수", "비율"])
    style_hdr(ws, 5, 3)
    for v in VERDICT_ORDER:
        c = stats.get(v, 0)
        ws.append([v, c, f"{c/total*100:.1f}%"])
        if v in fills:
            ws.cell(row=ws.max_row, column=1).fill = fills[v]
    ws.append(["합계", total, "100%"])
    ws.append([])
    # 원천별 교차
    ws.append(["원천별 교차"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    sources = sorted(set(r["원천시스템"] for r in audit_rows))
    ws.append(["원천"] + VERDICT_ORDER + ["합계"])
    style_hdr(ws, ws.max_row, len(VERDICT_ORDER) + 2)
    for s in sources:
        srows = [r for r in audit_rows if r["원천시스템"] == s]
        ws.append([s] + [sum(1 for r in srows if r["판정"] == v) for v in VERDICT_ORDER] + [len(srows)])
    for i, w in enumerate([26] + [14] * (len(VERDICT_ORDER) + 1), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 01_하드코딩(최우선)
    ws2 = wb.create_sheet("01_하드코딩(최우선)")
    ws2["A1"] = "GOLD 설계O·값 미주입 — DDL 무변경·SQL 배선만으로 해소 가능 (최우선 조치군)"
    ws2["A1"].font = title_font
    ws2.append([])
    ws2.append(["GOLD_모델", "GOLD_컬럼", "행", "하드코딩 패턴", "타 모델 실적재"])
    style_hdr(ws2, 3, 5)
    for (model, col), i in sorted(hardcoded.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        others = sorted({m for (m, c) in real if c == col and m != model})
        ws2.append([model, col, i["line"], i["pattern"], ", ".join(others) if others else "—"])
    for i, w in enumerate([34, 24, 7, 46, 34], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"

    # 02_계보매핑
    ws3 = wb.create_sheet("02_계보매핑")
    ws3["A1"] = "BRONZE→GOLD 명시 계보 (LINEAGE_MAP) — 개명 적재 탐지 근거"
    ws3["A1"].font = title_font
    ws3.append([])
    ws3.append(["BRONZE_테이블", "BRONZE_컬럼", "GOLD_컬럼", "GOLD_모델", "상태"])
    style_hdr(ws3, 3, 5)
    for (btbl, bcol), (gcol, gmodel) in sorted(LINEAGE_MAP.items()):
        st = "⚠️ 하드코딩 — 배선 필요" if (gmodel, gcol) in hardcoded else "✅ 적재"
        ws3.append([btbl, bcol, gcol, gmodel, st])
    for i, w in enumerate([26, 24, 22, 30, 24], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A4"

    # 03_원천별 상세
    header = ["BRONZE_테이블", "BRONZE_컬럼", "데이터타입", "판정", "신뢰도", "비고"]
    for src in sources:
        srows = [r for r in audit_rows if r["원천시스템"] == src]
        wss = wb.create_sheet(f"03_{src}")
        wss["A1"] = f"{src} — {len(srows)}컬럼"
        wss["A1"].font = title_font
        wss.append([])
        wss.append(header)
        style_hdr(wss, 3, len(header))
        for r in srows:
            wss.append([r[h] for h in header])
        for i, w in enumerate([30, 30, 14, 22, 20, 70], 1):
            wss.column_dimensions[get_column_letter(i)].width = w
        for row_idx in range(4, wss.max_row + 1):
            cell = wss.cell(row=row_idx, column=4)
            if cell.value in fills:
                cell.fill = fills[cell.value]
            for c in range(1, len(header) + 1):
                wss.cell(row=row_idx, column=c).alignment = wrap
                wss.cell(row=row_idx, column=c).border = border
        wss.freeze_panes = "A4"
        wss.auto_filter.ref = f"A3:{get_column_letter(len(header))}{wss.max_row}"

    # workspace FUSE 마운트는 zipfile 의 in-place seek 을 지원하지 않는다.
    # → 메모리(BytesIO)에서 zip 을 완성한 뒤 단일 바이너리 write 로 저장.
    import io
    buf = io.BytesIO()
    wb.save(buf)
    with open(os.path.join(OUT_DIR, BASENAME + ".xlsx"), "wb") as f:
        f.write(buf.getvalue())


def write_csv(audit_rows):
    path = os.path.join(OUT_DIR, BASENAME + ".csv")
    header = ["원천시스템", "BRONZE_테이블", "BRONZE_컬럼", "데이터타입", "판정", "신뢰도", "비고"]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([f"# 생성기: {GEN_PATH} | 러너: {RUNNER_PATH} | 감사일: {AUDIT_DATE}"])
        w.writerow([f"# {PROV}"])
        w.writerow([])
        w.writerow(header)
        for r in audit_rows:
            w.writerow([r[h] for h in header])


def write_md(audit_rows, stats, hardcoded, real):
    total = len(audit_rows)
    L = []
    L.append("<!-- LLM-METADATA")
    L.append("doc_id: BRONZE_EXPOSURE_AUDIT")
    L.append("doc_role: BRONZE 전 원천 전면 노출감사 — GOLD 도달 여부 판정 정본")
    L.append("project: GN_DW")
    L.append(f"audit_date: {AUDIT_DATE}")
    L.append(f"generator: {GEN_PATH}")
    L.append(f"runner: {RUNNER_PATH}")
    L.append("principle: P13(커버리지≠정확도)·P14(부재판정은 실측필수)")
    L.append("END-METADATA -->")
    L.append("")
    L.append("# BRONZE 노출감사 (전 원천 전면)")
    L.append("")
    L.append(f"> ⚙️ **자동 생성물** — 생성기 `{GEN_PATH}` / 러너 `{RUNNER_PATH}`. 직접 편집 금지.")
    L.append(f"> **감사일** {AUDIT_DATE} · **범위** BRONZE 전 원천 {total}컬럼 (CRM·AGENCY·ERP·GA4)")
    L.append("> **목적** \"보여줄 수 있는 BRONZE 데이터는 다 보여준다\" 충족 여부 실측")
    L.append("")
    L.append("## 0. 판정 기준 및 한계 (필독)")
    L.append("")
    L.append("| 판정 | 의미 | 조치 |")
    L.append("|---|---|---|")
    L.append("| 노출됨(GOLD) | GOLD 컬럼으로 실적재 (동명 또는 계보 매핑 확인) | 없음 |")
    L.append("| 대체노출(파생) | BRONZE 원본 대신 타 컬럼 파생값 적재(의도된 설계) | 없음 |")
    L.append("| ⚠️설계O·값미주입 | **GOLD에 컬럼 자리는 있으나 `0`/`CAST(NULL)` 하드코딩** | **최우선 배선** |")
    L.append("| SILVER까지만 | SILVER 적재 완료·GOLD 미승격 | 지표 수요 확인 후 승격 |")
    L.append("| 판정보류(동명이의) | 일반명이 타 계보 GOLD 컬럼과 충돌 — 근거 불충분 | 실측 확인(P14) |")
    L.append("| 미노출(검토대상) | 어느 층에도 이름이 없음 | 개명·VARIANT 승격 여부 개별 확인 |")
    L.append("| 제외(PII·본문·메타) | PII·본문·감사메타 — 의도적 비노출 | 없음 |")
    L.append("")
    L.append("**P13 한계 명시**: 본 감사는 *컬럼명 문자열 대조* 기반이다. ")
    L.append("개명 적재(예: `BRDC_DATE`→`BROADCAST_DATE`)는 `LINEAGE_MAP`에 등록된 건만 탐지된다. ")
    L.append("따라서 **'미노출' 판정은 부재 확정이 아니다** — 신뢰도 열이 `낮음(...)`인 행은 ")
    L.append("개별 실측(P14) 없이 '원천 부재'로 단정해서는 안 된다.")
    L.append("")
    L.append("**오탐 방지 설계**: 단순 전역 이름매칭은 무관한 동명 컬럼(예: AGENCY `YEAR` ↔ ")
    L.append("`DIM_DATE.YEAR`)을 '노출됨'으로 오판해 **실제 결손을 은폐**한다. 이를 막기 위해 ")
    L.append("(a) 계보 매핑은 **GOLD 모델까지 특정**하고, (b) 일반명은 `판정보류(동명이의)`로 격리하며, ")
    L.append("(c) 동명 컬럼이 여러 모델에 있을 때 **모델별 하드코딩/실적재를 구분**한다. ")
    L.append("예: `AD_COST` 는 FACT_AD_PERFORMANCE 실적재·FACT_BUDGET 하드코딩으로 상태가 다르다.")
    L.append("")
    L.append("## 1. 요약")
    L.append("")
    L.append("| 판정 | 건수 | 비율 |")
    L.append("|---|---|---|")
    for v in VERDICT_ORDER:
        c = stats.get(v, 0)
        L.append(f"| {v} | {c} | {c/total*100:.1f}% |")
    L.append(f"| **합계** | **{total}** | 100% |")
    L.append("")
    # 원천별 교차표
    L.append("### 원천별 교차")
    L.append("")
    sources = sorted(set(r["원천시스템"] for r in audit_rows))
    L.append("| 원천 | " + " | ".join(VERDICT_ORDER) + " | 합계 |")
    L.append("|---" * (len(VERDICT_ORDER) + 2) + "|")
    for s in sources:
        srows = [r for r in audit_rows if r["원천시스템"] == s]
        cells = [str(sum(1 for r in srows if r["판정"] == v)) for v in VERDICT_ORDER]
        L.append(f"| {s} | " + " | ".join(cells) + f" | {len(srows)} |")
    L.append("")
    L.append("## 2. ⚠️ 최우선 조치군 — GOLD 설계O·값 미주입")
    L.append("")
    L.append("dbt GOLD 모델에서 `0 as X_SK` 또는 `CAST(NULL AS ..) as X` 로 하드코딩된 컬럼 전량.")
    L.append("**GOLD 스키마에 자리가 이미 있으므로 DDL 변경 없이 SQL 배선만으로 해소 가능**한 군이다.")
    L.append("")
    L.append("| GOLD 컬럼 | 모델:행 | 하드코딩 패턴 | 타 모델 실적재 |")
    L.append("|---|---|---|---|")
    for (model, col), i in sorted(hardcoded.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        others = sorted({m for (m, c) in real if c == col and m != model})
        oth = ", ".join(others) if others else "—"
        L.append(f"| `{col}` | `{model}:{i['line']}` | `{i['pattern']}` | {oth} |")
    L.append("")
    L.append("> '타 모델 실적재'가 있는 행은 **해당 모델에서만** 결손이다. ")
    L.append("> 컬럼명만 보고 전역 해소로 오판하면 안 된다.")
    L.append("")
    L.append("### 2-1. BRONZE 실존 확인된 하드코딩 (즉시 배선 대상)")
    L.append("")
    L.append("`LINEAGE_MAP` 으로 BRONZE 원천이 확인된 건 — dbt 주석의 '원천 부재' 단정은 ")
    L.append("**오류이며 정정 대상**이다(P14 위반 사례).")
    L.append("")
    L.append("| BRONZE 테이블 | BRONZE 컬럼 | GOLD 컬럼 | GOLD 모델 | 상태 |")
    L.append("|---|---|---|---|---|")
    for (btbl, bcol), (gcol, gmodel) in sorted(LINEAGE_MAP.items()):
        if (gmodel, gcol) in hardcoded:
            L.append(f"| `{btbl}` | `{bcol}` | `{gcol}` | `{gmodel}` | ⚠️ 하드코딩 — 배선 필요 |")
    L.append("")
    L.append("## 3. 원천별 컬럼 상세")
    L.append("")
    for src in sources:
        srows = [r for r in audit_rows if r["원천시스템"] == src]
        L.append(f"### {src} ({len(srows)}컬럼)")
        L.append("")
        for tbl in sorted(set(r["BRONZE_테이블"] for r in srows)):
            trows = [r for r in srows if r["BRONZE_테이블"] == tbl]
            n_gold = sum(1 for r in trows if r["판정"] == "노출됨(GOLD)")
            n_hard = sum(1 for r in trows if r["판정"] == "⚠️설계O·값미주입")
            L.append(f"<details><summary><b>{tbl}</b> — {len(trows)}컬럼 "
                     f"(GOLD {n_gold} · 하드코딩 {n_hard})</summary>")
            L.append("")
            L.append("| 컬럼 | 타입 | 판정 | 신뢰도 | 비고 |")
            L.append("|---|---|---|---|---|")
            for r in trows:
                L.append(f"| `{r['BRONZE_컬럼']}` | {r['데이터타입']} | {r['판정']} "
                         f"| {r['신뢰도']} | {r['비고']} |")
            L.append("")
            L.append("</details>")
            L.append("")
    L.append("---")
    L.append("")
    L.append("## 관련 문서")
    L.append("")
    L.append("- `03_top-down_gold/05_필드 인벤토리.md` — 지표↔GOLD 필드 정본(P12)")
    L.append("- `03_top-down_gold/11_BRONZE적재 컬럼대조.md` — **CRM 전용·역방향**(원천요청서 대비 BRONZE 적재 확인). "
             "본 감사는 **전 원천·순방향**(BRONZE→GOLD 노출)으로 범위·방향이 다르며 상호 보완 관계.")
    L.append("- `20_issue/10_진단_원인분석.md` §8-I — 본 감사 기반 진단")
    L.append("")
    L.append(f"_감사일 {AUDIT_DATE} · Co-authored with CoCo_")
    with open(os.path.join(OUT_DIR, BASENAME + ".md"), "w", encoding="utf-8") as f:
        f.write("\n".join(L))


if __name__ == "__main__":
    main()
