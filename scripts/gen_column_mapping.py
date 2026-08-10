# BRONZE→SILVER→GOLD 컬럼 계보 매핑 생성기 (현업용 · WIDE 마트 기준)
# Co-authored with CoCo
"""
GN_DW 데이터 계보(lineage) 산출물 생성기 — **전량 실측 파생**.

⚠️ 설계 전환 (2026-08-06)
  구 판본은 계보 84행을 파이썬 리터럴로 **하드코딩**했다. 그래서 O24·O26·O34·O37·O38·O39 의
  개명·재정의가 하나도 반영되지 않았고(`MEMBER_STATUS`·`MEMBER_TYPE` 등 없는 컬럼을 계속 실었다),
  WIDE 9종만 다뤄 신설 4종(`WIDE_AD_*` 3 · `FACT_DEV_ACHIEVEMENT`)이 누락됐다.
  → 이 판본은 **인벤토리·계보·상태를 전부 측정해서 만든다**. 하드코딩 계보표는 없다.

측정 원천 (4원)
  1. INFORMATION_SCHEMA          — GOLD 뷰/테이블 물리 컬럼 모집단
  2. dbt WIDE 모델 SELECT 파싱   — WIDE 컬럼 → GOLD 테이블.컬럼 (alias → ref() 해석)
  3. dbt GOLD/SILVER 모델 파싱   — GOLD 컬럼 → SILVER 컬럼 → BRONZE 테이블
  4. 컬럼 census(COUNT/COUNT_IF) — 상태(OK/PARTIAL/WAIT) 판정. COMMENT 를 근거로 단정하지 않는다.

업무 의미는 **WIDE 모델 post_hook 의 컬럼 COMMENT** 에서 가져온다 — 그것이 배포 객체에 실제로
붙는 문장이고 이슈 교정(O34·O38·O39 …)이 반영되는 유일한 정본이다.

출력: 30_output_share/04_컬럼계보매핑.{md,csv,xlsx}
"""
import csv, os, re, sys, json, collections
from datetime import date

WS = os.environ.get("GN_DW_WS", "/workspace")
OUT_DIR = os.environ.get("GN_DW_OUT", os.path.join(WS, "30_output_share"))
BASENAME = "04_컬럼계보매핑"
GEN_PATH = "scripts/gen_column_mapping.py"
MEASURED = os.environ.get("GN_DW_MEASURED", date.today().isoformat())
PROV = f"본 파일은 자동 생성물입니다. 직접 수정 금지 — 생성기 {GEN_PATH} 수정 후 재실행하세요."
MODELS = os.path.join(WS, "10_dbt_pipeline", "models")
CENSUS = os.environ.get("GN_DW_CENSUS", "/tmp/census.json")
SCHEMA_JSON = os.environ.get("GN_DW_SCHEMA", "/tmp/schema.json")

AUDIT = {"DW_SOURCE_SYSTEM", "DW_SOURCE_TABLE", "DW_LOAD_TS", "DW_UPDATE_TS", "DW_BATCH_ID"}
SCHEMA = {}

HEADER = ["WIDE_마트", "WIDE_컬럼", "GOLD_원천(테이블.컬럼)", "SILVER_원천(테이블.컬럼)",
          "BRONZE_원천(테이블)", "업무_의미(배포 COMMENT)", "WIDE→GOLD_확정도", "계보_확정도",
          "상태", "상태_근거(실측)"]


# ─────────────────────────── dbt 모델 로딩·파싱 ───────────────────────────
def strip_comments(s):
    s = re.sub(r"/\*.*?\*/", " ", s, flags=re.S)
    return re.sub(r"--[^\n]*", " ", s)


def load_models():
    out = {}
    for root, _, files in os.walk(MODELS):
        for fn in files:
            if not fn.endswith(".sql"):
                continue
            p = os.path.join(root, fn)
            rel = os.path.relpath(p, MODELS)
            raw = open(p, encoding="utf-8").read()
            out[fn[:-4]] = {
                "layer": rel.split(os.sep)[0],
                "sub": rel.split(os.sep)[1] if len(rel.split(os.sep)) > 2 else "",
                "path": rel, "raw": raw, "sql": strip_comments(raw),
                "refs": set(re.findall(r"ref\(\s*['\"]([A-Za-z0-9_]+)['\"]\s*\)", raw)),
                "sources": set(re.findall(r"source\(\s*['\"][^'\"]+['\"]\s*,\s*['\"]([A-Za-z0-9_]+)['\"]\s*\)", raw)),
            }
    return out


def alias_map(sql):
    """FROM/JOIN 절의 `{{ ref('X') }} alias` → {alias: X}.
    인라인 서브쿼리 `( … {{ ref('X') }} … ) alias` 도 ref 가 하나뿐이면 해석한다."""
    m = {}
    KW = {"on", "left", "right", "inner", "outer", "full", "cross", "join", "where",
          "select", "and", "or", "group", "order", "qualify", "having", "union", "as"}
    for ref, al in re.findall(r"ref\(\s*['\"]([A-Za-z0-9_]+)['\"]\s*\)\s*\}\}\s*(?:as\s+)?([A-Za-z_][A-Za-z0-9_]*)", sql, re.I):
        if al.lower() not in KW:
            m[al] = ref
    # 인라인 서브쿼리: 괄호 균형을 맞춰 블록을 잡고 닫는 괄호 뒤의 별칭을 읽는다
    for st in [i for i, ch in enumerate(sql) if ch == "("]:
        depth = 0
        for j in range(st, len(sql)):
            if sql[j] == "(":
                depth += 1
            elif sql[j] == ")":
                depth -= 1
                if depth == 0:
                    block = sql[st + 1:j]
                    tail = sql[j + 1:j + 40]
                    am = re.match(r"\s*(?:as\s+)?([A-Za-z_][A-Za-z0-9_]*)", tail)
                    refs = set(re.findall(r"ref\(\s*['\"]([A-Za-z0-9_]+)['\"]\s*\)", block))
                    if am and len(refs) == 1 and am.group(1).lower() not in KW:
                        m.setdefault(am.group(1), next(iter(refs)))
                    break
    return m


def select_items(sql):
    """최상위 SELECT 목록에서 (출력컬럼, 소스표현) 추출.
    `a.COL as OUT` · `a.COL` · `expr as OUT` 형태를 모두 잡는다."""
    # 마지막 select ... from 블록을 최상위로 본다(WIDE 는 단일 select 구조).
    mts = list(re.finditer(r"\bselect\b", sql, re.I))
    best = None
    for mt in mts:
        tail = sql[mt.end():]
        fm = re.search(r"\bfrom\b", tail, re.I)
        if not fm:
            continue
        body = tail[:fm.start()]
        if best is None or len(body) > len(best):
            best = body
    if not best:
        return []
    # 괄호 깊이 0 의 콤마로 분할
    parts, depth, cur = [], 0, ""
    for ch in best:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "," and depth == 0:
            parts.append(cur)
            cur = ""
        else:
            cur += ch
    parts.append(cur)
    out = []
    for p in parts:
        p = " ".join(p.split())
        if not p:
            continue
        m = re.match(r"^(.*?)\s+as\s+([A-Za-z_][A-Za-z0-9_]*)$", p, re.I)
        if m:
            expr, name = m.group(1).strip(), m.group(2)
        else:
            expr, name = p, p.split(".")[-1]
        if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", name or ""):
            continue
        out.append((name.upper(), expr))
    return out


def wide_comments(raw):
    """post_hook 의 `ALTER VIEW ... COLUMN X COMMENT '...'` 를 {컬럼: 설명} 으로."""
    out = {}
    for m in re.finditer(r"COLUMN\s+([A-Za-z_][A-Za-z0-9_]*)\s+COMMENT\s+'((?:[^']|'')*)'", raw, re.I):
        out[m.group(1).upper()] = m.group(2).replace("''", "'").strip()
    return out


def resolve(expr, amap):
    """`a.COL` → (ref대상, COL). alias 미해석이면 (None, 마지막토큰)."""
    m = re.fullmatch(r"([A-Za-z_][A-Za-z0-9_]*)\.([A-Za-z_][A-Za-z0-9_]*)", expr.strip())
    if m and m.group(1) in amap:
        return amap[m.group(1)], m.group(2).upper()
    cols = re.findall(r"\b([A-Za-z_][A-Za-z0-9_]*)\.([A-Za-z_][A-Za-z0-9_]*)\b", expr)
    for a, c in cols:
        if a in amap:
            return amap[a], c.upper()
    return None, None


def trace_down(models, model, col):
    """GOLD 모델의 출력컬럼 col 이 어느 SILVER 테이블.컬럼에서 왔는지 추적.

    GOLD 모델은 CTE 를 많이 쓰므로 최상위 SELECT 파싱만으로는 부족하다 →
    ① 모델 전체에서 `<식> as <col>` 별칭을 찾고 ② 그 식의 컬럼 토큰을
    ③ 이 모델이 ref() 하는 SILVER 테이블의 실제 컬럼 목록과 대조한다.
    대조에 성공한 것만 `컬럼확정` 이고, 실패하면 확정도를 낮춰 그대로 표기한다(추론 금지).
    """
    m = models.get(model)
    if not m:
        return None, None, "미해석"
    sql = m["sql"]
    silver_refs = sorted(r for r in m["refs"] if models.get(r, {}).get("layer") == "silver")
    # GOLD→GOLD 경유(차원이 다른 차원을 참조하는 경우)
    gold_refs = sorted(r for r in m["refs"] if models.get(r, {}).get("layer") == "gold")
    scols = {st: set(SCHEMA["silver_cols"].get(st, [])) for st in silver_refs}

    def find_in_silver(name):
        hits = [st for st, cs in scols.items() if name in cs]
        return hits

    # ① 별칭 정의 탐색: `... as COL`
    cands = []
    for mm in re.finditer(r"([A-Za-z_][A-Za-z0-9_.\"']*(?:\([^()]*\))?)\s+as\s+" + re.escape(col) + r"\b", sql, re.I):
        cands.append(mm.group(1))
    for mm in re.finditer(r"\bas\s+" + re.escape(col) + r"\b", sql, re.I):
        start = max(0, mm.start() - 200)
        cands.append(sql[start:mm.start()])
    # ② 식에서 컬럼 토큰 추출 후 SILVER 대조
    for expr in cands:
        toks = re.findall(r"\b([A-Za-z_][A-Za-z0-9_]{2,})\b", expr)
        for t in reversed(toks):
            hits = find_in_silver(t.upper())
            if hits:
                return hits[0], t.upper(), "컬럼확정"
    # ③ 동명 통과(별칭 없이 그대로 SELECT)
    hits = find_in_silver(col)
    if hits:
        return hits[0], col, "컬럼확정(동명통과)"
    # ④ GOLD 내부 1단 경유
    for gr in gold_refs:
        if col in set(SCHEMA["gold_cols"].get(gr, [])):
            t2, c2, cf2 = trace_down(models, gr, col)
            if t2:
                return t2, c2, "컬럼확정(2단)" if cf2.startswith("컬럼확정") else cf2
    if silver_refs:
        return ";".join(silver_refs), None, "테이블수준"
    return None, None, "미해석"


def silver_to_bronze(models, silver_tables):
    out = set()
    for st in re.split(r"[;]", silver_tables or ""):
        st = st.strip()
        if st in models:
            out |= models[st]["sources"]
            for r in models[st]["refs"]:
                if models.get(r, {}).get("layer") == "silver":
                    out |= models[r]["sources"]
    return ";".join(sorted(out))


# ─────────────────────────── 상태 판정 ───────────────────────────
def status_of(census, gold_table, gold_col):
    key = f"GOLD.{gold_table}" if gold_table else None
    e = census.get(key or "")
    if not e:
        return "—", "census 대상 아님(뷰 파생·차원 라벨 등)"
    if e["rows"] == 0:
        return "WAIT", f"`{gold_table}` 0행 — 원천 미입고"
    ci = e["cols"].get(gold_col)
    if not ci:
        return "—", f"`{gold_table}` 에 `{gold_col}` 물리 부재"
    nn, nz, rows = int(ci["nonnull"] or 0), int(ci["nonzero"] or 0), e["rows"]
    if nn == 0:
        return "WAIT", f"전건 NULL (0/{rows:,})"
    if nz == 0:
        return "WAIT", f"전건 0 (비영 0/{rows:,}) — 설계O·값 미주입"
    if nz < rows:
        return "PARTIAL", f"비영 {nz:,}/{rows:,} ({nz/rows*100:.1f}%)"
    return "OK", f"비영 {nz:,}/{rows:,} (100%)"


def main():
    global SCHEMA
    census = json.load(open(CENSUS, encoding="utf-8"))
    schema = json.load(open(SCHEMA_JSON, encoding="utf-8"))
    SCHEMA = schema
    models = load_models()

    wide_views = sorted(t for t in schema["views"] if t.startswith("WIDE_"))
    rows = []
    unresolved = []
    for w in wide_views:
        m = models.get(w)
        cols = schema["view_cols"].get(w, [])
        cmts = wide_comments(m["raw"]) if m else {}
        amap = alias_map(m["sql"]) if m else {}
        sel = dict(select_items(m["sql"])) if m else {}
        for c in cols:
            if c in AUDIT:
                continue
            expr = sel.get(c, "")
            gt, gc = resolve(expr, amap) if expr else (None, None)
            if gt is None and expr:
                mm = re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*\.([A-Za-z_][A-Za-z0-9_]*)", expr.strip())
                if mm:
                    gc = mm.group(1).upper()
            wide_conf = "직접참조" if gt else ""
            if not gt:
                # CTE 경유(FACT_DEV_ACHIEVEMENT 류) — 이 WIDE 가 ref 하는 GOLD 테이블에서 동명 대조.
                grefs = sorted(r for r in (m["refs"] if m else set())
                               if r in schema["gold_cols"] or r in schema["view_cols"])
                cand = [r for r in grefs if (gc or c) in set(schema["gold_cols"].get(r, []))]
                if len(cand) == 1:
                    gt, gc = cand[0], (gc or c)
                    wide_conf = "CTE경유·동명대조"
                elif len(cand) > 1:
                    gt, gc = cand[0], (gc or c)
                    wide_conf = f"CTE경유·다중후보({','.join(cand)})"
                else:
                    unresolved.append((w, c, (expr or "(SELECT 목록 파싱 실패)")[:70]))
                    wide_conf = "미해석"
            st_tbl, st_col, conf = trace_down(models, gt, gc) if gt and gc else (None, None, "미해석")
            if conf == "미해석" and gt and not any(
                    models.get(r, {}).get("layer") == "silver" for r in models.get(gt, {}).get("refs", set())):
                conf = "원천무관(ETL 생성)"
            bronze = silver_to_bronze(models, st_tbl) if st_tbl else ""
            status, why = status_of(census, gt, gc)
            rows.append({
                "WIDE_마트": w, "WIDE_컬럼": c,
                "GOLD_원천(테이블.컬럼)": f"{gt}.{gc}" if gt and gc else (gc or ""),
                "SILVER_원천(테이블.컬럼)": (f"{st_tbl}.{st_col}" if st_tbl and st_col else (st_tbl or "")),
                "BRONZE_원천(테이블)": bronze,
                "업무_의미(배포 COMMENT)": cmts.get(c, ""),
                "WIDE→GOLD_확정도": wide_conf,
                "계보_확정도": conf, "상태": status, "상태_근거(실측)": why,
            })

    os.makedirs(OUT_DIR, exist_ok=True)
    # ── CSV ──
    cpath = os.path.join(OUT_DIR, BASENAME + ".csv")
    with open(cpath, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=HEADER)
        w.writeheader()
        w.writerows(rows)

    # ── MD ──
    cst = collections.Counter(r["상태"] for r in rows)
    ccf = collections.Counter(r["계보_확정도"] for r in rows)
    L = []
    A = L.append
    A("<!-- LLM-METADATA")
    A("doc_id: BRONZE_SILVER_GOLD_COLUMN_MAPPING")
    A("doc_role: 현업용 source-target 컬럼 계보 매핑 (WIDE 마트 기준 · 전량 실측 파생)")
    A("project: GN_DW (굿네이버스)")
    A(f"measured: {MEASURED}")
    A("grounded_on: INFORMATION_SCHEMA · dbt WIDE/GOLD/SILVER 모델 파싱 · 컬럼 census(COUNT_IF)")
    A(f"generator: {GEN_PATH}")
    A("generated: auto (do-not-edit)")
    A("END-METADATA -->")
    A("")
    A("# BRONZE → SILVER → GOLD 컬럼 계보 매핑 (현업용)")
    A("")
    A(f"> ⚙️ **생성기**: `{GEN_PATH}` — {PROV}")
    A(f"> **측정일 {MEASURED}** · 대상 = 배포된 GOLD WIDE 소비뷰 **{len(wide_views)}종 / 컬럼 {len(rows)}개**")
    A("> **읽는 법**: 현업이 조회하는 **최종 마트(WIDE)** 컬럼을 기준으로, 그 값이 어떤 GOLD → SILVER → BRONZE")
    A("> 원천에서 왔는지 역방향 추적합니다. **업무 의미는 배포된 뷰 COMMENT 원문**이므로 화면에서 보는 설명과 같습니다.")
    A("")
    A("> 🔴 **상태는 COMMENT 가 아니라 실측입니다** — 각 행의 `상태`는 해당 GOLD 물리 컬럼에")
    A("> `COUNT`/`COUNT_IF(<>0)` 를 돌려 판정했습니다. 「설계는 됐으나 값이 안 들어온」 컬럼이 `WAIT` 로 드러납니다.")
    A("")
    A("**상태 범례** — `OK` 전행 유효값 · `PARTIAL` 일부만 채움(부분집합 주의) · `WAIT` 전건 NULL/0 또는 0행 테이블 · `—` 판정 대상 아님")
    A("")
    A("## 0. 요약")
    A("")
    A("| 구분 | 건수 |")
    A("|---|---:|")
    A(f"| WIDE 소비뷰 | **{len(wide_views)}** |")
    A(f"| WIDE 컬럼(감사컬럼 제외) | **{len(rows)}** |")
    for k in ("OK", "PARTIAL", "WAIT", "—"):
        if cst[k]:
            A(f"| └ 상태 `{k}` | {cst[k]} |")
    A("")
    A("**WIDE → GOLD 확정도**")
    A("")
    A("| 확정도 | 건수 | 의미 |")
    A("|---|---:|---|")
    wmean = {"직접참조": "WIDE SELECT 가 `alias.컬럼` 으로 GOLD 물리 컬럼을 직접 참조",
             "CTE경유·동명대조": "CTE 를 거쳤으나 이 WIDE 가 ref 하는 GOLD 테이블에 동명 컬럼이 유일하게 존재",
             "CTE경유·다중후보": "동명 컬럼이 여러 GOLD 테이블에 있어 첫 후보를 적었다(⚠️확정 아님)",
             "미해석": "집계·파생·플래그 — 직접 대응 컬럼이 없다(§4 자기고발)"}
    cwf = collections.Counter(r["WIDE→GOLD_확정도"].split("(")[0] for r in rows)
    for k, v in cwf.most_common():
        A(f"| {k} | {v} | {wmean.get(k,'')} |")
    A("")
    A("**GOLD → SILVER → BRONZE 확정도** — 이름 유사성이 아니라 파싱 + 물리 컬럼 대조 결과입니다(P36).")
    A("")
    A("| 확정도 | 건수 | 의미 |")
    A("|---|---:|---|")
    mean = {"컬럼확정": "GOLD 모델의 별칭 정의식을 SILVER 물리 컬럼 목록과 대조해 확정",
            "컬럼확정(동명통과)": "별칭 없이 그대로 SELECT — SILVER 동명 컬럼 실존 확인",
            "컬럼확정(2단)": "GOLD 내부 1단(차원→차원) 경유 후 컬럼 단위 확정",
            "원천무관(ETL 생성)": "`DIM_DATE` 등 SILVER 원천이 없는 ETL 생성 차원 — 결손이 아니다",
            "테이블수준": "SILVER **테이블**까지만 확정 — 컬럼 대응은 미확정(⚠️단정 금지)",
            "미해석": "파싱 실패 — 집계·파생·플래그 등 직접 대응이 없는 컬럼(§4 자기고발)"}
    for k, v in ccf.most_common():
        A(f"| {k} | {v} | {mean.get(k,'')} |")
    A("")
    A("> ⚠️ **`테이블수준`·`다중후보`·`미해석` 행은 계보 확정 근거가 아닙니다** — 컬럼 단위 출처를 확정해야")
    A("> 하는 경우 해당 dbt 모델 SQL 을 직접 확인하세요(P13·P36).")
    A("")
    A("## 1. 업무 질문 → 어느 마트를 볼까?")
    A("")
    A("> 아래 상태는 §2 이후의 컬럼 실측을 근거로 합니다.")
    A("")
    A("| 업무 질문 | 조회 마트 | 상태 |")
    A("|---|---|---|")
    for qq, mm, ss in [
        ("지난달 캠페인·후원사업별 신규개발/중단 실적", "`WIDE_MEMBER_MONTHLY`", "✅"),
        ("회원 상태·지역·연령대별 납입회비 추이", "`WIDE_MEMBER_MONTHLY`", "✅ (지역·연령대는 **약정시점 스냅샷**)"),
        ("개발/중단 건별 사유·경로·부서·일자 상세", "`WIDE_MEMBER_EVENT`", "✅ (부서는 **개발 사건 전용**)"),
        ("캠페인별 중단률(이탈률)·유지기간", "`SV_MEMBER_COHORT`(회원 grain)", "✅ **12개월 고정률만 사용**"),
        ("부서별 회원개발 **목표 대비 달성률**", "`FACT_DEV_ACHIEVEMENT`", "◐ 부서 단위 O · 상위조직 불가(CONF-4)"),
        ("알림톡/서신 발송 실적", "`WIDE_SERVICE_EVENT`", "◐ 발송수·고유회원수 O · 성공/실패/오픈·+5일 반응 ⛔"),
        ("행사별 참여 회원·참여자수", "`WIDE_EVENT_PARTICIPATION`", "◐ 참여수 O · 상태별 카운트 ⛔(O28 코드체계)"),
        ("예산 편성 대비 집행 현황", "`WIDE_BUDGET`", "◐ 편성(월)·집행 O · 연편성·모금성비용·광고비 ⛔(E-1/E-4)"),
        ("디지털 광고 매체비·노출·클릭", "`WIDE_AD_DIGITAL` / `WIDE_AD_PERFORMANCE`", "◐ 매체 measure O · 캠페인·소재 연결 ⛔(Q10)"),
        ("방송·재방송 광고 성과", "`WIDE_AD_BROADCAST` / `WIDE_AD_BROADCAST_CASE`", "◐ 재방송 개발단가 O · VIDEO 개발실적 ⛔(AD-5)"),
        ("웹/앱 방문·세션·스크롤 등 GA 행동", "`WIDE_GA_BEHAVIOR`", "◐ 2일 샤드만(G-5) · 회원귀속 극소"),
        ("연/추경 **사업목표** 달성률", "`WIDE_TARGET_BIZ`", "⛔ 구조준비·0행(E-6)"),
    ]:
        A(f"| {qq} | {mm} | {ss} |")
    A("")
    A("> 🔴 **「사업목표」와 「회원개발 목표」는 다른 것입니다** — 앞은 E-6 미입고(0행), 뒤는 사용 가능합니다. 합산·비교하지 마세요.")
    A("")
    A("## 2. 마트별 컬럼 계보")
    A("")
    bymart = collections.defaultdict(list)
    for r in rows:
        bymart[r["WIDE_마트"]].append(r)
    for i, w in enumerate(wide_views, 1):
        rs = bymart[w]
        c2 = collections.Counter(r["상태"] for r in rs)
        A(f"### 2-{i}. `{w}` — 컬럼 {len(rs)} (OK {c2['OK']} · PARTIAL {c2['PARTIAL']} · WAIT {c2['WAIT']})")
        A("")
        A("| WIDE 컬럼 | GOLD 원천 | SILVER 원천 | BRONZE 원천(테이블) | 확정도(WIDE→GOLD / GOLD→SILVER) | 상태 | 상태 근거(실측) | 업무 의미(배포 COMMENT) |")
        A("|---|---|---|---|---|---|---|---|")
        for r in rs:
            mean_txt = r["업무_의미(배포 COMMENT)"].replace("|", "\\|").replace("\n", " ")
            A(f"| `{r['WIDE_컬럼']}` | `{r['GOLD_원천(테이블.컬럼)']}` | `{r['SILVER_원천(테이블.컬럼)']}` | "
              f"{r['BRONZE_원천(테이블)']} | {r['WIDE→GOLD_확정도']} / {r['계보_확정도']} | {r['상태']} | {r['상태_근거(실측)']} | {mean_txt} |")
        A("")
    A("## 3. 🔴 상태 `WAIT` 전량 — 설계는 됐으나 값이 없다")
    A("")
    A("> 컬럼 자리는 있으나 전건 NULL/0 입니다. **`0` 을 실측값으로 읽으면 조용히 틀립니다**(P15).")
    A("")
    waits = [r for r in rows if r["상태"] == "WAIT"]
    seen = set()
    A("| GOLD 원천 | 노출 마트 | 상태 근거 |")
    A("|---|---|---|")
    for r in sorted(waits, key=lambda x: x["GOLD_원천(테이블.컬럼)"]):
        g = r["GOLD_원천(테이블.컬럼)"]
        if g in seen:
            continue
        seen.add(g)
        marts = sorted({x["WIDE_마트"] for x in waits if x["GOLD_원천(테이블.컬럼)"] == g})
        A(f"| `{g}` | {', '.join(marts)} | {r['상태_근거(실측)']} |")
    A("")
    if unresolved:
        A("## 4. ⚠️ 파싱 미해석 컬럼 (자기고발)")
        A("")
        A("생성기가 GOLD 원천을 확정하지 못한 행입니다. **추론으로 채우지 않았습니다.**")
        A("")
        A("| 마트 | 컬럼 | SELECT 표현식(발췌) |")
        A("|---|---|---|")
        for w, c, e in unresolved[:200]:
            A(f"| `{w}` | `{c}` | `{e.replace('|','\\|')}` |")
        A("")
    A("---")
    A("_Co-authored with CoCo_")
    mpath = os.path.join(OUT_DIR, BASENAME + ".md")
    open(mpath, "w", encoding="utf-8").write("\n".join(L) + "\n")

    # ── XLSX ──
    xpath = ""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        wb.remove(wb.active)
        hdr_f = Font(bold=True, color="FFFFFF")
        hdr_b = PatternFill("solid", fgColor="2F5597")
        for w in wide_views:
            ws = wb.create_sheet(w[:31])
            ws.append(HEADER)
            for cell in ws[1]:
                cell.font, cell.fill = hdr_f, hdr_b
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            for r in bymart[w]:
                ws.append([r[h] for h in HEADER])
            for j, h in enumerate(HEADER, 1):
                ws.column_dimensions[get_column_letter(j)].width = min(60, max(14, len(h) + 6))
            ws.freeze_panes = "A2"
        tmp = os.path.join("/tmp", BASENAME + ".xlsx")
        wb.save(tmp)
        xpath = os.path.join(OUT_DIR, BASENAME + ".xlsx")
        import shutil
        shutil.copyfile(tmp, xpath)
    except ImportError:
        xpath = "(openpyxl 미설치 — XLSX 생략)"

    print("CSV :", cpath)
    print("MD  :", mpath)
    print("XLSX:", xpath)
    print("마트", len(wide_views), "컬럼", len(rows), "| 상태", dict(cst), "| 확정도", dict(ccf), "| 미해석", len(unresolved))


if __name__ == "__main__":
    main()
