# BRONZE→SILVER→GOLD source-target 컬럼 매핑 생성기 (현업용, WIDE 마트 기준 시트 분할)
# Co-authored with CoCo
"""
GN_DW 데이터 계보(lineage) 매핑 산출물 생성기.
근거(정본):
  - 03_top-down_gold/08_silver의존.md   (SILVER→GOLD 컬럼 계보)
  - 10_dbt_pipeline/models/silver/_sources.yml (BRONZE→SILVER 원천 테이블)
  - silver/gold dbt 모델 SELECT 식 (컬럼 변환 규칙)
출력: MD(가독) · CSV(가공) · XLSX(현업 공유, WIDE 시트 분할)
"""
import csv, os

OUT_DIR = os.path.realpath("/workspace/03_top-down_gold")
BASENAME = "34_BRONZE-SILVER-GOLD_컬럼매핑_현업"

# ── 컬럼: 마트, GOLD컬럼, 업무의미, SILVER(테이블.컬럼), BRONZE(테이블.컬럼), 변환규칙, 상태
# 상태: OK=적재완료 · PARTIAL=일부대기 · WAIT=원천대기
HEADER = ["WIDE_마트", "GOLD_컬럼", "업무_의미", "SILVER_원천(테이블.컬럼)",
          "BRONZE_원천(테이블.컬럼)", "변환_규칙", "상태"]

ROWS = [
    # ─────────────────────────── WIDE_MEMBER_MONTHLY ───────────────────────────
    ["WIDE_MEMBER_MONTHLY", "MONTH_KEY", "실적 월(YYYYMM)", "CRM_MEMBER_DEV.OCCRRNC_DE 등 거래일", "TM_MM_FDRM_MBER_DVLP_AMT.OCCRRNC_DE", "거래일→월 롤업(YYYYMM)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "MEMBER_DK", "회원 식별키", "CRM_MEMBER.MEMBER_DK", "TM_MM_FDRM_MBER_INFO.MBER_NO / TM_MM_ONCE_MBER_INFO.ONCE_MBER_NO", "정기∪일시 회원번호 통합키", "OK"],
    ["WIDE_MEMBER_MONTHLY", "DEV_MEMBERS", "개발(명)", "CRM_MEMBER_DEV(MBER_NO distinct)", "TM_MM_FDRM_MBER_DVLP_AMT.MBER_NO", "월×회원 COUNT(DISTINCT)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "DEV_CNT", "개발(건)", "CRM_MEMBER_DEV.SPNSR_AMT", "TM_MM_FDRM_MBER_DVLP_AMT.SPNSR_AMT", "SUM(금액)/10000 basis", "OK"],
    ["WIDE_MEMBER_MONTHLY", "INCREASE_MEMBERS", "증액(명)", "CRM_MEMBER_AMT_CHANGE(RDCAMT_YN='N')", "TM_MM_FDRM_MBER_IRSD.SPNSR_AMT·RDCAMT_YN", "증액분 회원 COUNT", "OK"],
    ["WIDE_MEMBER_MONTHLY", "DECREASE_CNT", "감액(건)", "CRM_MEMBER_AMT_CHANGE(RDCAMT_YN='Y')", "TM_MM_FDRM_MBER_IRSD.SPNSR_AMT·RDCAMT_YN", "SUM(감액금액)/10000", "OK"],
    ["WIDE_MEMBER_MONTHLY", "CHURN_CNT", "이탈(건)=취소+감액", "CRM_MEMBER_DISCONTINUE + CRM_MEMBER_AMT_CHANGE", "TM_MM_FDRM_MBER_SPNSR_DSCNTC + TM_MM_FDRM_MBER_IRSD", "SUM(취소+감액)/10000 (신규 표준지표)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "ACTIVE_MEMBERS", "활동회원(명)", "CRM_MEMBER_STATUS_HIST + CRM_MEMBER_SPONSOR_BIZ", "TH_MM_FDRM_MBER_STNG_DTLS + TM_MM_FDRM_MBER_SPNSR_BSNS", "월말 시점 스냅샷 COUNT", "OK"],
    ["WIDE_MEMBER_MONTHLY", "UNPAID_CNT", "미납(건)", "CRM_PAYMENT_BILLING(PAY_STAT_CD=F∪NULL)", "TM_PM_MBRFEE_ACMSLT.PAY_STAT_CD", "미납 상태 건수(C-3 정의)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "STOP_CNT", "중단(건)", "CRM_MEMBER_DISCONTINUE", "TM_MM_FDRM_MBER_SPNSR_DSCNTC.SPNSR_DSCNTC_DE", "월 중단 건수", "OK"],
    ["WIDE_MEMBER_MONTHLY", "REGULAR_FEE", "정기회비(원)", "CRM_PAYMENT_BILLING.PAY_AMT", "TM_PM_MBRFEE_ACMSLT.PAY_AMT", "정기 납입액 SUM(원단위)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "PAID_FEE", "납입회비(원)", "CRM_PAYMENT_BILLING.PAY_AMT", "TM_PM_MBRFEE_ACMSLT.PAY_AMT + TM_PM_DNTN_DTLS.PAY_AMT", "회비∪기부금 납입 SUM", "OK"],
    ["WIDE_MEMBER_MONTHLY", "BILLED_AMT", "청구(원)", "CRM_PAYMENT_BILLING.RQEST_AMT", "TM_PM_MBRFEE_ACMSLT.RQEST_AMT", "청구액 SUM(행기준)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "REDONATE_FLAG", "재후원 여부", "CRM_MEMBER_RESPONSOR", "TM_MM_FDRM_MBER_RE_SPNSR", "재후원 이력 존재 플래그", "OK"],
    ["WIDE_MEMBER_MONTHLY", "INBOUND_CALL_CNT", "인바운드콜수", "(CRM 부재)", "(원천 없음)", "현업 수기입력(비-CRM, C-8)", "WAIT"],
    ["WIDE_MEMBER_MONTHLY", "TS_CALL_CNT", "TS콜수", "(CRM 부재)", "(원천 없음)", "현업 수기입력(비-CRM, C-8)", "WAIT"],
    ["WIDE_MEMBER_MONTHLY", "MEMBER_GENDER", "성별", "CRM_MEMBER.SEX", "TM_MM_FDRM_MBER_INFO.SEX", "1/3→M, 2/4→F, NULL→NULL 정규화", "OK"],
    ["WIDE_MEMBER_MONTHLY", "MEMBER_REGION", "지역", "CRM_MEMBER_DEV.AREA_CD(스냅샷)", "TM_MM_FDRM_MBER_DVLP_AMT.AREA_CD", "CM018 코드→라벨(개발시점 스냅샷)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "MEMBER_AGE_BAND", "연령대", "CRM_MEMBER_DEV.AGE(스냅샷)", "TM_MM_FDRM_MBER_DVLP_AMT.AGE", "연령 코드화(생년 raw 미적재)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "MEMBER_STATUS", "회원상태", "CRM_MEMBER.MBER_STAT_CD + STATUS_HIST", "TM_MM_FDRM_MBER_INFO.MBER_STAT_CD", "MM010 코드→라벨 + SCD2 이력", "OK"],
    ["WIDE_MEMBER_MONTHLY", "MEMBER_TYPE", "회원구분(정기/일시)", "CRM_MEMBER.MEMBER_TYPE(파생)", "TM_MM_FDRM_MBER_INFO / TM_MM_ONCE_MBER_INFO", "FDRM/ONCE 크로스워크(테이블 출처 기반)", "OK"],
    ["WIDE_MEMBER_MONTHLY", "CAMPAIGN_BRAND", "브랜드", "CRM_CAMPAIGN.BRND_NM", "TM_CM_BRND_MNG.BRND_NM", "캠페인→브랜드 LEFT JOIN", "OK"],
    ["WIDE_MEMBER_MONTHLY", "CAMPAIGN_NAME", "캠페인명", "CRM_CAMPAIGN.CMPGN_NM", "TM_CM_CMPGN_MNG.CMPGN_NM", "-", "OK"],
    ["WIDE_MEMBER_MONTHLY", "SPONSORSHIP_NAME", "후원사업명", "CRM_SPONSORSHIP.SPNSR_BSNS_NM", "TM_CM_SPNSR_BSNS_INFO.SPNSR_BSNS_NM", "-", "OK"],
    ["WIDE_MEMBER_MONTHLY", "PAYMENT_METHOD", "결제수단", "CRM_PAYMENT_METHOD.CARD_DIV_CD", "TM_PM_SETLE_INFO.CARD_DIV_CD", "코드→라벨", "OK"],

    # ─────────────────────────── WIDE_SERVICE_EVENT ───────────────────────────
    ["WIDE_SERVICE_EVENT", "DATE_SK", "발송일(YYYYMMDD)", "CRM_SEND_REQUEST 발송일", "TM_MS_EMAIL_SNDNG / TM_MS_MSG_AT_SNDNG / TM_MS_PSTMTR_SNDNG", "채널별 발송일 통합", "OK"],
    ["WIDE_SERVICE_EVENT", "MEMBER_DK", "발송 대상 회원", "CRM_SEND_MEMBER.MBER_NO", "TD_MS_EMAIL_SNDNG_DTLS / TD_MS_MSG_AT_SNDNG_DTLS / TD_MS_PSTMTR_SNDNG_DTL", "발송상세 회원키", "OK"],
    ["WIDE_SERVICE_EVENT", "SEND_MEMBERS", "발송수(명)", "CRM_SEND_MEMBER(MBER_NO distinct)", "TD_MS_*_SNDNG_DTLS", "COUNT(DISTINCT 회원)", "OK"],
    ["WIDE_SERVICE_EVENT", "SUCCESS_MEMBERS", "성공수(명)", "CRM_SEND_RESULT(성공)", "TD_MS_EMAIL_LQY_SNDNG / TD_MS_MSG_AT_LQY_SNDNG / TD_MS_PSTMTR_LQY_SNDNG", "성과 성공건 회원 COUNT", "OK"],
    ["WIDE_SERVICE_EVENT", "FAIL_MEMBERS", "실패수(명)", "CRM_SEND_RESULT(실패)", "TD_MS_*_LQY_SNDNG", "성과 실패건 회원 COUNT", "OK"],
    ["WIDE_SERVICE_EVENT", "LETTER_PART_MEMBERS", "서신참여(명)", "CRM_RELATION_ACTIVITY(LETTER)", "TM_RM_RELATNSP_LETTER_INFO", "발송×서신 매칭", "OK"],
    ["WIDE_SERVICE_EVENT", "GIFT_PART_AMT", "선물금참여(원)", "CRM_RELATION_ACTIVITY.GFTMNEY", "TM_RM_RELATNSP_GFTMNEY_INFO.GFTMNEY", "선물금 SUM(원단위)", "OK"],
    ["WIDE_SERVICE_EVENT", "D5_LETTER_PART_MEMBERS", "+5일 서신참여(명)", "CRM_RELATION_ACTIVITY × CRM_SEND_*", "TM_RM_RELATNSP_LETTER_INFO", "발송일+5일 윈도우 매칭", "OK"],
    ["WIDE_SERVICE_EVENT", "D5_INCREASE_PART_MEMBERS", "+5일 증액참여(명)", "CRM_MEMBER_AMT_CHANGE × CRM_SEND_*", "TM_MM_FDRM_MBER_IRSD", "발송일+5일 윈도우 매칭", "OK"],
    ["WIDE_SERVICE_EVENT", "SEND_TITLE", "발송 제목", "CRM_SEND_REQUEST.TIT", "SND_REQ_MST.SEND_TITLE / TM_MS_*_SNDNG", "-", "OK"],
    ["WIDE_SERVICE_EVENT", "SEND_STATUS", "발송상태", "CRM_SEND_MEMBER.SNDNG_RST_CD", "TD_MS_*_SNDNG_DTLS", "코드→라벨", "OK"],
    ["WIDE_SERVICE_EVENT", "APP_PUSH_SEND_CNT", "앱푸시 발송수", "(ADMIN 제외)", "(원천 미채택)", "어드민 제외 확정(2026-07-09)→미채움", "WAIT"],
    ["WIDE_SERVICE_EVENT", "SERVICE_CHANNEL", "발송 채널", "CRM_SEND_REQUEST.SEND_CHANNEL", "SND_REQ_MST / TM_MS_*_SNDNG", "이메일/문자/우편 채널", "OK"],

    # ─────────────────────────── WIDE_AD_PERFORMANCE ───────────────────────────
    ["WIDE_AD_PERFORMANCE", "PERF_DATE_SK", "광고 실적일", "AGENCY_AD_PERFORMANCE 실적일", "DGT/REBRDC/VIDEO_AD_CMPGN_DTLS", "유형별 실적일 통합", "PARTIAL"],
    ["WIDE_AD_PERFORMANCE", "AD_COST", "광고비(원)", "AGENCY_AD_PERFORMANCE 광고비", "DGT/REBRDC/VIDEO_AD_CMPGN_DTLS(광고비 컬럼 3종)", "유형별 비용 정제→UNION", "OK"],
    ["WIDE_AD_PERFORMANCE", "IMPRESSIONS", "노출수", "AGENCY_AD_PERFORMANCE", "DGT_AD_CMPGN_DTLS(노출)", "디지털(DGT)만 존재", "PARTIAL"],
    ["WIDE_AD_PERFORMANCE", "CLICKS", "클릭수", "AGENCY_AD_PERFORMANCE", "DGT_AD_CMPGN_DTLS(클릭)", "디지털(DGT)만 존재", "PARTIAL"],
    ["WIDE_AD_PERFORMANCE", "INBOUND_CALL", "인입콜수", "AGENCY_AD_PERFORMANCE", "REBRDC/VIDEO_AD_CMPGN_DTLS", "TRY_TO_NUMBER 캐스팅", "PARTIAL"],
    ["WIDE_AD_PERFORMANCE", "GA_CONV_MEMBERS", "GA전환수(명)", "GA4_EVENT(전환 이벤트)", "BRONZE_GA4.events_YYYYMMDD", "전환정의(O5) 회신 대기", "WAIT"],
    ["WIDE_AD_PERFORMANCE", "CAMPAIGN_BRAND", "브랜드", "CRM_CAMPAIGN.BRND_NM", "TM_CM_BRND_MNG.BRND_NM", "캠페인 이름매칭(Q10) 대기", "WAIT"],
    ["WIDE_AD_PERFORMANCE", "AD_MEDIA_NAME", "매체명", "AGENCY_AD_CREATIVE.MEDIA_NAME", "DGT_AD_CMPGN_DTLS.MEDIA_NM 등", "유형별 산재→UNION", "PARTIAL"],
    ["WIDE_AD_PERFORMANCE", "AD_PLATFORM", "플랫폼", "AGENCY_AD_CREATIVE.PLATFORM", "DGT/REBRDC/VIDEO_AD_CMPGN_DTLS", "-", "PARTIAL"],
    ["WIDE_AD_PERFORMANCE", "DEVICE_TYPE", "디바이스", "GA4_DEVICE.DEVICE_TYPE", "BRONZE_GA4.events(device)", "PC/M/APP", "PARTIAL"],

    # ─────────────────────────── WIDE_MEMBER_EVENT ───────────────────────────
    ["WIDE_MEMBER_EVENT", "DATE_SK", "거래일", "CRM_MEMBER_DEV.OCCRRNC_DE", "TM_MM_FDRM_MBER_DVLP_AMT.OCCRRNC_DE", "-", "OK"],
    ["WIDE_MEMBER_EVENT", "MEMBER_DK", "회원 식별키", "CRM_MEMBER.MEMBER_DK", "TM_MM_FDRM_MBER_INFO.MBER_NO", "-", "OK"],
    ["WIDE_MEMBER_EVENT", "DEV_CNT/STOP_CNT", "개발/중단 건·명", "CRM_MEMBER_DEV + CRM_MEMBER_DISCONTINUE", "TM_MM_FDRM_MBER_DVLP_AMT + TM_MM_FDRM_MBER_SPNSR_DSCNTC", "개발·중단 이벤트 건수", "OK"],
    ["WIDE_MEMBER_EVENT", "STOP_REASON", "중단 사유", "CRM_MEMBER_DISCONTINUE.DSCNTC_RSN_CD", "TM_MM_FDRM_MBER_SPNSR_DSCNTC.DSCNTC_RSN_CD", "사유코드→라벨", "OK"],
    ["WIDE_MEMBER_EVENT", "STOP_CHANNEL", "중단 경로", "CRM_MEMBER_DISCONTINUE.DSCNTC_PATH", "TM_MM_FDRM_MBER_SPNSR_DSCNTC.DSCNTC_PATH", "-", "OK"],

    # ─────────────────────────── WIDE_EVENT_PARTICIPATION ───────────────────────────
    ["WIDE_EVENT_PARTICIPATION", "DATE_SK", "행사일", "CRM_EVENT.STRT_DE", "TM_MS_EVENT.STRT_DE / TM_MS_CRMN", "행사∪캠페인행사", "OK"],
    ["WIDE_EVENT_PARTICIPATION", "MEMBER_DK", "참여 회원", "CRM_EVENT_PARTICIPATION.MBER_NO", "TD_MS_EVENT_PRTCPNT_DTL / TD_MS_CRMN_PRTCPNT", "-", "OK"],
    ["WIDE_EVENT_PARTICIPATION", "EVENT_NAME", "행사명", "CRM_EVENT.EVENT_NM", "TM_MS_EVENT.EVENT_NM", "-", "OK"],
    ["WIDE_EVENT_PARTICIPATION", "PARTICIPATION_STATUS", "참여상태", "CRM_EVENT_PARTICIPATION.PARTCPT_STAT_CD", "TD_MS_EVENT_PRTCPNT_DTL.PARTCPT_STAT_CD", "모집/참여/취소/당첨", "OK"],
    ["WIDE_EVENT_PARTICIPATION", "RECEIPT_AMT", "참여 납입액", "CRM_EVENT_PARTICIPATION.RCPMNY_AMT", "TD_MS_EVENT_PRTCPNT_DTL.RCPMNY_AMT", "원단위", "OK"],
    ["WIDE_EVENT_PARTICIPATION", "VIEW_CNT", "행사 조회수", "(ADMIN 제외)", "(원천 미채택)", "어드민 제외 확정→미채움", "WAIT"],

    # ─────────────────────────── WIDE_TARGET_DEV ───────────────────────────
    ["WIDE_TARGET_DEV", "MONTH_KEY", "목표 월(YYYYMM)", "CRM_DEV_TARGET(STDYY+STDR_MT)", "TM_CM_MBER_DVLP_GOAL.STDYY·STDR_MT", "연+월→YYYYMM", "OK"],
    ["WIDE_TARGET_DEV", "ORG_NM", "조직/부서", "CRM_ORG.DEPT_NM(←DEPT_ID)", "TM_CM_DEPT_INFO.DEPT_NM", "부서 계층 전개", "OK"],
    ["WIDE_TARGET_DEV", "DEV_TYPE", "개발구분", "CRM_DEV_TARGET.MBER_DVLP_DIV_CD", "TM_CM_MBER_DVLP_GOAL.MBER_DVLP_DIV_CD", "코드→라벨", "OK"],
    ["WIDE_TARGET_DEV", "GOAL_CNT", "목표 건수", "CRM_DEV_TARGET.GOAL_CNT", "TM_CM_MBER_DVLP_GOAL.GOAL_CNT", "-", "OK"],

    # ─────────────────────────── WIDE_BUDGET ───────────────────────────
    ["WIDE_BUDGET", "MONTH_KEY", "예산 월(YYYYMM)", "ERP_BUDGET.MONTH_KEY", "BDGT_ACMSLT_LEDGER(YEAR + 월컬럼)", "wide→long 언피벗(12개월)", "OK"],
    ["WIDE_BUDGET", "BUDGET_ITEM(계층)", "예산과목(장~세세목)", "ERP_BUDGET_ITEM(JANG~SUBDTL)", "BDGT_ACMSLT_LEDGER.JANG_NM~SUBDTL_ITEM_NM", "6단계 계층 보존·MD5 키", "OK"],
    ["WIDE_BUDGET", "YEAR_BUDGET_AMT", "편성예산(원)", "ERP_BUDGET.YEAR_BUDGET_AMT", "BDGT_ACMSLT_LEDGER.YEAR_BDGT_AMT_n", "월별 언피벗", "OK"],
    ["WIDE_BUDGET", "EXEC_AMT", "집행예산(원)", "ERP_BUDGET.EXEC_AMT", "BDGT_ACMSLT_LEDGER.EXEC_AMT_n", "월별 언피벗", "OK"],
    ["WIDE_BUDGET", "ORG_NM", "조직/부서", "ERP_BUDGET_ITEM.BDGT_UNIT_NM", "BDGT_ACMSLT_LEDGER.BDGT_UNIT_NM", "예산단위명(이름)", "OK"],
    ["WIDE_BUDGET", "FUNDRAISING_COST", "모금성비용(원)", "(원천 부재)", "(ERP 원장에 없음)", "원천 부재(E-1)→미채움", "WAIT"],
    ["WIDE_BUDGET", "AD_COST", "광고비(원)", "AGENCY 보강 대기", "DGT/REBRDC/VIDEO_AD_CMPGN_DTLS", "AGENCY_COST 보강(E-4)", "WAIT"],

    # ─────────────────────────── WIDE_GA_BEHAVIOR ───────────────────────────
    ["WIDE_GA_BEHAVIOR", "DATE_SK", "행동 일자", "GA4_EVENT.event_date", "BRONZE_GA4.events_YYYYMMDD.event_date", "샤드 UNION", "PARTIAL"],
    ["WIDE_GA_BEHAVIOR", "SESSION_CNT", "세션수", "GA4_EVENT.ga_session_id", "BRONZE_GA4.events(event_params: ga_session_id)", "param 승격·비가산 주의", "PARTIAL"],
    ["WIDE_GA_BEHAVIOR", "SCROLL_DEPTH", "스크롤 깊이", "GA4_EVENT.percent_scrolled", "BRONZE_GA4.events(event_params: percent_scrolled)", "param 승격", "PARTIAL"],
    ["WIDE_GA_BEHAVIOR", "PAGE_PATH", "페이지 경로", "GA4_EVENT.page_location", "BRONZE_GA4.events(event_params: page_location)", "degenerate", "PARTIAL"],
    ["WIDE_GA_BEHAVIOR", "IDENTITY(회원귀속)", "GA↔회원 연결", "GA4_IDENTITY.member_id", "BRONZE_GA4.events(user_id / user_pseudo_id)", "user_id→session_fill(브리지, S)", "WAIT"],
    ["WIDE_GA_BEHAVIOR", "UTM_SOURCE/MEDIUM", "유입 소스/매체", "GA4_TRAFFIC_SOURCE.UTM_*", "BRONZE_GA4.events(traffic_source)", "-", "PARTIAL"],
]

# DIM/FACT 부록(파워유저용) — 08_silver의존.md 요약
DIM_FACT_ROWS = [
    ["DIM", "DIM_MEMBER", "회원 마스터(SCD2)", "CRM_MEMBER (+ STATUS_HIST)", "TM_MM_FDRM_MBER_INFO / TM_MM_ONCE_MBER_INFO / TH_MM_FDRM_MBER_STNG_DTLS", "정기∪일시 통합·상태이력", "OK"],
    ["DIM", "DIM_MEMBER_IDENTITY", "신원 브리지(GA↔회원)", "CRM_MEMBER + GA4_IDENTITY", "TM_MM_FDRM_MBER_INFO + BRONZE_GA4.events", "user_id 매칭(0행·비활성)", "WAIT"],
    ["DIM", "DIM_CAMPAIGN", "캠페인", "CRM_CAMPAIGN", "TM_CM_CMPGN_MNG (+BRND_MNG+MKTNG_CMPGN_MNG)", "브랜드/마케팅캠페인 조인", "OK"],
    ["DIM", "DIM_SPONSORSHIP", "후원사업", "CRM_SPONSORSHIP", "TM_CM_SPNSR_BSNS_INFO", "-", "OK"],
    ["DIM", "DIM_ORG", "조직/부서", "CRM_ORG", "TM_CM_DEPT_INFO", "부서 계층 재귀전개", "OK"],
    ["DIM", "DIM_PAYMENT", "결제수단", "CRM_PAYMENT_METHOD", "TM_PM_SETLE_INFO", "-", "OK"],
    ["DIM", "DIM_REASON", "사유(미납/중단)", "CRM_MEMBER_DISCONTINUE (또는 CRM_CODE MM005)", "TM_MM_FDRM_MBER_SPNSR_DSCNTC", "사유코드→라벨", "OK"],
    ["DIM", "DIM_EVENT", "행사", "CRM_EVENT", "TM_MS_EVENT / TM_MS_CRMN", "행사∪캠페인행사", "OK"],
    ["DIM", "DIM_SERVICE", "발송구분", "CRM_SEND_REQUEST", "SND_REQ_MST (SEND_GBN_TOP/MID/BOT)", "대/중/소 코드→라벨", "OK"],
    ["DIM", "DIM_GA_SOURCE", "GA 트래픽소스", "GA4_TRAFFIC_SOURCE", "BRONZE_GA4.events(traffic_source)", "UTM 파싱", "PARTIAL"],
    ["DIM", "DIM_GA_EVENT", "GA 이벤트 분류", "GA4_EVENT_DIM", "BRONZE_GA4.events(event_params)", "category/label/action", "PARTIAL"],
    ["DIM", "DIM_DEVICE", "디바이스", "GA4_DEVICE", "BRONZE_GA4.events(device)", "PC/M/APP", "PARTIAL"],
    ["DIM", "DIM_AD_CREATIVE", "광고소재", "AGENCY_AD_CREATIVE", "DGT/REBRDC/VIDEO_AD_CMPGN_DTLS", "유형별 정제→UNION", "PARTIAL"],
    ["DIM", "DIM_BUDGET_ITEM", "예산과목 마스터", "ERP_BUDGET_ITEM", "BDGT_ACMSLT_LEDGER(장~세세목)", "DISTINCT 마스터·MD5 키", "OK"],
    ["DIM", "DIM_DATE", "날짜 차원", "(ETL 생성)", "(원천 무관)", "팩트 일자범위로 생성", "OK"],
    ["FACT", "FACT_MEMBER_MONTHLY", "회원 월 팩트", "CRM_MEMBER_* 다수", "TM_MM_FDRM_* / TM_PM_*", "월 grain 롤업", "OK"],
    ["FACT", "FACT_MEMBER_EVENT", "회원 이벤트 팩트", "CRM_MEMBER_DEV + DISCONTINUE", "TM_MM_FDRM_MBER_DVLP_AMT + _SPNSR_DSCNTC", "일 grain", "OK"],
    ["FACT", "FACT_SERVICE_EVENT", "발송 팩트", "CRM_SEND_* + RELATION_ACTIVITY", "TM_MS_*_SNDNG(+DTLS/LQY) + TM_RM_*", "발송+반응(+5일)", "OK"],
    ["FACT", "FACT_EVENT_PARTICIPATION", "행사참여 팩트", "CRM_EVENT + PARTICIPATION", "TM_MS_EVENT/CRMN + TD_MS_*_PRTCPNT*", "-", "OK"],
    ["FACT", "FACT_TARGET_DEV", "개발목표 팩트", "CRM_DEV_TARGET", "TM_CM_MBER_DVLP_GOAL", "월×조직×개발구분", "OK"],
    ["FACT", "FACT_GA_BEHAVIOR", "GA행동 팩트", "GA4_EVENT + IDENTITY", "BRONZE_GA4.events_YYYYMMDD", "샤드 통합", "PARTIAL"],
    ["FACT", "FACT_AD_PERFORMANCE", "광고성과 팩트", "AGENCY_AD_PERFORMANCE + GA4_EVENT", "DGT/REBRDC/VIDEO_AD_CMPGN_DTLS + events", "캠페인매칭(Q10)·전환(O5) 대기", "PARTIAL"],
    ["FACT", "FACT_BUDGET", "예산 팩트", "ERP_BUDGET (+AGENCY 보강)", "BDGT_ACMSLT_LEDGER", "편성/집행 O·모금성비용 대기", "PARTIAL"],
    ["FACT", "FACT_TARGET_BIZ", "사업목표 팩트", "(원천 부재)", "(ERP 원장≠사업목표)", "별도 입고 대기(E-6)", "WAIT"],
]

# 업무 질문 → 마트 인덱스
INDEX_ROWS = [
    ["지난달 캠페인·후원사업별 신규개발/이탈 실적", "WIDE_MEMBER_MONTHLY", "✅"],
    ["회원 상태·지역·연령대별 활동회원·납입회비", "WIDE_MEMBER_MONTHLY", "✅"],
    ["개발/중단 건별 사유·경로 상세", "WIDE_MEMBER_EVENT", "✅"],
    ["알림톡/서신 발송 후 +5일 내 반응(증액·서신참여)", "WIDE_SERVICE_EVENT", "✅"],
    ["행사별 참여 회원·납입 실적", "WIDE_EVENT_PARTICIPATION", "✅"],
    ["조직·부서별 회원개발 목표 대비 달성률", "WIDE_TARGET_DEV", "✅"],
    ["예산 편성 대비 집행 현황(조직·예산과목별)", "WIDE_BUDGET", "◐ 모금성비용/광고비 대기"],
    ["디지털 광고 매체비·노출·클릭 성과", "WIDE_AD_PERFORMANCE", "◐ 캠페인연결/ROAS 대기"],
    ["웹/앱 방문·세션·스크롤 등 GA 행동", "WIDE_GA_BEHAVIOR", "◐ 회원귀속 대기"],
    ["광고 대비 CRM 후원 전환율(ROAS)", "WIDE_AD_PERFORMANCE+신원브리지", "🔜 로드맵"],
    ["연/추경 사업목표 대비 달성률", "(FACT_TARGET_BIZ)", "⛔ 원천 대기(E-6)"],
]

def write_csv():
    path = os.path.join(OUT_DIR, BASENAME + ".csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADER)
        w.writerows(ROWS)
        w.writerow([])
        w.writerow(["# 부록: DIM/FACT 계층 매핑 (파워유저용)"])
        w.writerow(["계층", "GOLD_객체", "업무_의미", "SILVER_원천", "BRONZE_원천", "변환_규칙", "상태"])
        w.writerows(DIM_FACT_ROWS)
    return path

def write_md():
    path = os.path.join(OUT_DIR, BASENAME + ".md")
    marts = []
    for r in ROWS:
        if r[0] not in marts:
            marts.append(r[0])
    lines = []
    lines.append("<!-- LLM-METADATA")
    lines.append("doc_id: BRONZE_SILVER_GOLD_COLUMN_MAPPING")
    lines.append("doc_role: 현업용 source-target 컬럼 계보 매핑 (WIDE 마트 기준)")
    lines.append("project: GN_DW (굿네이버스)")
    lines.append("created: 2026-07-15")
    lines.append("grounded_on: 08_silver의존.md · _sources.yml · silver/gold dbt models")
    lines.append("END-METADATA -->")
    lines.append("")
    lines.append("# BRONZE → SILVER → GOLD 컬럼 매핑 (현업용)")
    lines.append("")
    lines.append("> **읽는 법**: 현업이 조회하는 **최종 결과(GOLD/WIDE 마트)** 컬럼을 기준으로,")
    lines.append("> 그 값이 어떤 **SILVER(정제)** → **BRONZE(원천)** 컬럼에서 왔는지 역방향으로 추적합니다.")
    lines.append("> 상태: **OK** 사용가능 · **PARTIAL** 일부 대기 · **WAIT** 원천 입고 대기")
    lines.append("")
    lines.append("## 0. 업무 질문 → 어느 마트를 볼까?")
    lines.append("")
    lines.append("| 업무 질문 | 조회 마트 | 상태 |")
    lines.append("|---|---|---|")
    for q, mart, st in INDEX_ROWS:
        lines.append(f"| {q} | `{mart}` | {st} |")
    lines.append("")
    for i, mart in enumerate(marts, 1):
        lines.append(f"## {i}. `{mart}`")
        lines.append("")
        lines.append("| GOLD 컬럼 | 업무 의미 | SILVER 원천 | BRONZE 원천 | 변환 규칙 | 상태 |")
        lines.append("|---|---|---|---|---|---|")
        for r in ROWS:
            if r[0] == mart:
                lines.append(f"| `{r[1]}` | {r[2]} | `{r[3]}` | `{r[4]}` | {r[5]} | {r[6]} |")
        lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 부록. DIM/FACT 계층 매핑 (파워유저·엔지니어용)")
    lines.append("")
    lines.append("| 계층 | GOLD 객체 | 업무 의미 | SILVER 원천 | BRONZE 원천 | 변환 규칙 | 상태 |")
    lines.append("|---|---|---|---|---|---|---|")
    for r in DIM_FACT_ROWS:
        lines.append(f"| {r[0]} | `{r[1]}` | {r[2]} | `{r[3]}` | `{r[4]}` | {r[5]} | {r[6]} |")
    lines.append("")
    lines.append("---")
    lines.append("_Co-authored with CoCo_")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path

def write_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=13, color="1F4E78")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_fill = {"OK": PatternFill("solid", fgColor="E2EFDA"),
                   "PARTIAL": PatternFill("solid", fgColor="FFF2CC"),
                   "WAIT": PatternFill("solid", fgColor="FCE4D6")}

    def style_header(ws, row_idx, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = wrap; cell.border = border

    # INDEX 시트
    ws = wb.active; ws.title = "00_INDEX"
    ws["A1"] = "GN_DW 데이터 매핑 — 업무 질문별 조회 마트 안내"; ws["A1"].font = title_font
    ws.append([]); ws.append(["업무 질문", "조회 마트", "상태"])
    style_header(ws, 3, 3)
    for q, mart, st in INDEX_ROWS:
        ws.append([q, mart, st])
    ws.column_dimensions["A"].width = 52; ws.column_dimensions["B"].width = 34; ws.column_dimensions["C"].width = 22
    for r in range(4, ws.max_row + 1):
        for c in range(1, 4):
            ws.cell(row=r, column=c).alignment = wrap; ws.cell(row=r, column=c).border = border
    ws.append([]); ws.append(["범례", "OK=사용가능 · PARTIAL/◐=일부대기 · WAIT/⛔/🔜=원천 입고 대기"])

    marts = []
    for r in ROWS:
        if r[0] not in marts: marts.append(r[0])

    def build_sheet(title, rows, is_dimfact=False):
        ws = wb.create_sheet(title[:31])
        ws["A1"] = title; ws["A1"].font = title_font
        head = (["계층", "GOLD 객체", "업무 의미", "SILVER 원천", "BRONZE 원천", "변환 규칙", "상태"]
                if is_dimfact else
                ["GOLD 컬럼", "업무 의미", "SILVER 원천(테이블.컬럼)", "BRONZE 원천(테이블.컬럼)", "변환 규칙", "상태"])
        ws.append([]); ws.append(head)
        style_header(ws, 3, len(head))
        for r in rows:
            ws.append(r if is_dimfact else r[1:])
        widths = ([10, 22, 22, 30, 40, 30, 12] if is_dimfact
                  else [26, 22, 34, 44, 34, 12])
        for i, wdt in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = wdt
        st_col = len(head)
        for r in range(4, ws.max_row + 1):
            for c in range(1, len(head) + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = wrap; cell.border = border
            stv = ws.cell(row=r, column=st_col).value
            if stv in status_fill:
                ws.cell(row=r, column=st_col).fill = status_fill[stv]
        ws.freeze_panes = "A4"
        return ws

    for mart in marts:
        rows = [r for r in ROWS if r[0] == mart]
        build_sheet(mart, rows, is_dimfact=False)

    build_sheet("_부록_DIM_FACT", DIM_FACT_ROWS, is_dimfact=True)

    # 스테이지 FUSE 마운트는 ZipFile 랜덤액세스 쓰기를 지원하지 않으므로 /tmp에 저장 후 복사
    import shutil
    tmp_path = os.path.join("/tmp", BASENAME + ".xlsx")
    wb.save(tmp_path)
    path = os.path.join(OUT_DIR, BASENAME + ".xlsx")
    shutil.copyfile(tmp_path, path)
    return path

if __name__ == "__main__":
    print("CSV :", write_csv())
    print("MD  :", write_md())
    try:
        print("XLSX:", write_xlsx())
    except ImportError:
        print("XLSX: openpyxl 미설치 — pip install 후 재실행 필요")
